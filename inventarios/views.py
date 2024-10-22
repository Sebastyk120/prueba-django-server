import io

from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test, login_required
from django.db import transaction
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.urls import reverse_lazy
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.generic.edit import CreateView, UpdateView
from django_tables2 import SingleTableView
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook import Workbook

from comercial.models import Referencias
from .forms import ItemForm, SearchForm, EditarItemForm, EliminarItemForm
from .models import Bodega, Item, Movimiento, Inventario
from .tables import MovimientoTable, ItemTable, InventarioTable


# Funciones para validar el Grupo del usuario y si puede acceder a la vista:

def es_miembro_del_grupo(nombre_grupo):
    def es_miembro(user):
        return user.groups.filter(name=nombre_grupo).exists()

    return es_miembro


# -------------------------------------- Exportar Items Etnico (Movimientos) Excel: ---------------------------------
@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Etnico'), login_url='home'))
def exportar_items_etnico(request):
    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Inventario de Items'
    font = Font(bold=True)
    fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Encabezados
    columns = ['Referencia', 'Cantidad Cajas', 'Tipo Documento', 'Documento', 'Bodega', 'Proveedor',
               'Fecha Movimiento', 'Propiedad', 'Observaciones', 'Usuario']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill

    # Obtener los datos de tu modelo
    queryset = Item.objects.filter(bodega__exportador__nombre='Etnico')

    # Agregar datos al libro de trabajo
    for row_num, item in enumerate(queryset, start=2):
        row = [
            item.numero_item.nombre,
            item.cantidad_cajas,
            item.get_tipo_documento_display(),
            item.documento,
            item.bodega.nombre,
            item.proveedor.nombre,
            item.fecha_movimiento.strftime("%Y-%m-%d"),
            item.propiedad.nombre,
            item.observaciones,
            item.user.username,
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventario_items_etnico.xlsx"'

    return response


# -------------------------------------- Exportar Items Fieldex (Movimientos) Excel: ---------------------------------
@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Fieldex'), login_url='home'))
def exportar_items_fieldex(request):
    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Inventario de Items'
    font = Font(bold=True)
    fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Encabezados
    columns = ['Referencia', 'Cantidad Cajas', 'Tipo Documento', 'Documento', 'Bodega', 'Proveedor',
               'Fecha Movimiento', 'Propiedad', 'Observaciones', 'Usuario']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill

    # Obtener los datos de tu modelo
    queryset = Item.objects.filter(bodega__exportador__nombre='Fieldex')

    # Agregar datos al libro de trabajo
    for row_num, item in enumerate(queryset, start=2):
        row = [
            item.numero_item.nombre,
            item.cantidad_cajas,
            item.get_tipo_documento_display(),  # RECORDAR QUE ES PARA LISTAS.
            item.documento,
            item.bodega.nombre,  #
            item.proveedor.nombre,  #
            item.fecha_movimiento.strftime("%Y-%m-%d"),  # RECORDAR FORMATEO.
            item.propiedad.nombre,  #
            item.observaciones,
            item.user.username,  # Nombre de Usuario.
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventario_items_fieldex.xlsx"'

    return response


# -------------------------------------- Exportar Items Juan Matas (Movimientos) Excel: ---------------------------------
@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Juan_Matas'), login_url='home'))
def exportar_items_juan(request):
    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Inventario de Juan Matas'
    font = Font(bold=True)
    fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Encabezados
    columns = ['Referencia', 'Cantidad Cajas', 'Tipo Documento', 'Documento', 'Bodega', 'Proveedor',
               'Fecha Movimiento', 'Propiedad', 'Observaciones', 'Usuario']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill

    # Obtener los datos de tu modelo
    queryset = Item.objects.filter(bodega__exportador__nombre='Juan_Matas')

    # Agregar datos al libro de trabajo
    for row_num, item in enumerate(queryset, start=2):
        row = [
            item.numero_item.nombre,
            item.cantidad_cajas,
            item.get_tipo_documento_display(),  # RECORDAR QUE ES PARA LISTAS.
            item.documento,
            item.bodega.nombre,  #
            item.proveedor.nombre,  #
            item.fecha_movimiento.strftime("%Y-%m-%d"),  # RECORDAR FORMATEO.
            item.propiedad.nombre,  #
            item.observaciones,
            item.user.username,  # Nombre de Usuario.
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventario_items_juan.xlsx"'

    return response


# ------------------ Exportacion de inventarios  Excel General --------------------------------------------------------
@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Heavens'), login_url='home'))
def exportar_inventario_excel(request):
    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Inventario General'
    font = Font(bold=True)
    fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Encabezados
    columns = ['Referencia', 'Exportador', 'Compras Efectivas', 'Saldos Iniciales', 'Salidas', 'Traslado Propio',
               'Traslado Remisionado', 'Ventas', 'Venta Contenedor', 'Stock Actual']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill

    # Obtener los datos de tu modelo
    queryset = Inventario.objects.all()

    # Agregar datos al libro de trabajo
    for row_num, item in enumerate(queryset, start=2):
        row = [
            item.numero_item.nombre,
            item.numero_item.exportador.nombre,
            item.compras_efectivas,
            item.saldos_iniciales,
            item.salidas,
            item.traslado_propio,
            item.traslado_remisionado,
            item.ventas,
            item.venta_contenedor,
            # Calculo stock actual
            (item.compras_efectivas + item.saldos_iniciales) - (
                    item.salidas + item.traslado_propio + item.traslado_remisionado + item.ventas)
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventario_general.xlsx"'

    return response


# ------------------ Exportacion de inventarios  Excel Etnico --------------------------------------------------------
@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Etnico'), login_url='home'))
def exportar_inventario_etnico(request):
    # Libro De Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Inventario Etnico'
    font = Font(bold=True)
    fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Encabezados
    columns = ['Referencia', 'Exportador', 'Compras Efectivas', 'Saldos Iniciales', 'Salidas', 'Traslado Propio',
               'Traslado Remisionado', 'Ventas', 'Venta Contenedor', 'Stock Actual']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill

    # Filtrar datos Etnico
    queryset = Inventario.objects.filter(numero_item__exportador__nombre='Etnico')

    # Filas para Etnico
    for row_num, item in enumerate(queryset, start=2):
        row = [
            item.numero_item.nombre,
            item.numero_item.exportador.nombre,
            item.compras_efectivas,
            item.saldos_iniciales,
            item.salidas,
            item.traslado_propio,
            item.traslado_remisionado,
            item.ventas,
            item.venta_contenedor,
            # Calculo Stock Actual
            (item.compras_efectivas + item.saldos_iniciales) - (
                    item.salidas + item.traslado_propio + item.traslado_remisionado + item.ventas)
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventario_etnico.xlsx"'

    return response


# ------------------ Exportacion de inventarios  Excel Fieldex --------------------------------------------------------
@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Fieldex'), login_url='home'))
def exportar_inventario_fieldex(request):
    # Libro De Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Inventario Fieldex'
    font = Font(bold=True)
    fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Encabezados
    columns = ['Referencia', 'Exportador', 'Compras Efectivas', 'Saldos Iniciales', 'Salidas', 'Traslado Propio',
               'Traslado Remisionado', 'Ventas', 'Venta Contenedor', 'Stock Actual']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill

    # Filtrar datos Fieldex
    queryset = Inventario.objects.filter(numero_item__exportador__nombre='Fieldex')

    # Filas para Etnico
    for row_num, item in enumerate(queryset, start=2):
        row = [
            item.numero_item.nombre,
            item.numero_item.exportador.nombre,
            item.compras_efectivas,
            item.saldos_iniciales,
            item.salidas,
            item.traslado_propio,
            item.traslado_remisionado,
            item.ventas,
            item.venta_contenedor,
            # Calculo Stock Actual
            (item.compras_efectivas + item.saldos_iniciales) - (
                    item.salidas + item.traslado_propio + item.traslado_remisionado + item.ventas)
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventario_fieldex.xlsx"'

    return response


# ------------------ Exportacion de inventarios Excel Juan Matas-----------------------------------------------------
@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Juan_Matas'), login_url='home'))
def exportar_inventario_juan(request):
    # Libro De Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Inventario Juan Matas'
    font = Font(bold=True)
    fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Encabezados
    columns = ['Referencia', 'Exportador', 'Compras Efectivas', 'Saldos Iniciales', 'Salidas', 'Traslado Propio',
               'Traslado Remisionado', 'Ventas', 'Venta Contenedor', 'Stock Actual']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill

    # Filtrar datos Juan Matas
    queryset = Inventario.objects.filter(numero_item__exportador__nombre='Juan_Matas')

    # Filas para Etnico
    for row_num, item in enumerate(queryset, start=2):
        row = [
            item.numero_item.nombre,
            item.numero_item.exportador.nombre,
            item.compras_efectivas,
            item.saldos_iniciales,
            item.salidas,
            item.traslado_propio,
            item.traslado_remisionado,
            item.ventas,
            item.venta_contenedor,
            # Calculo Stock Actual
            (item.compras_efectivas + item.saldos_iniciales) - (
                    item.salidas + item.traslado_propio + item.traslado_remisionado + item.ventas)
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventario_juan_matas.xlsx"'

    return response


# ------------------- Lista de Items Etnico  --------------------------------------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Etnico'), login_url=reverse_lazy('home')), name='dispatch')
class ItemListView(SingleTableView):
    model = Item
    table_class = ItemTable
    template_name = 'recibo_items_list_etnico.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(bodega__exportador__nombre='Etnico')
        form = self.form_class(self.request.GET)
        if form.is_valid() and form.cleaned_data.get('item_busqueda'):
            item_busqueda = form.cleaned_data.get('item_busqueda')
            queryset = queryset.filter(numero_item__nombre__icontains=item_busqueda)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# -------------------------------  Formulario - Crear Item Etnico - Modal (Inventario Real) ----------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Etnico'), login_url=reverse_lazy('home')), name='dispatch')
class ItemCreateView(CreateView):
    model = Item
    form_class = ItemForm
    template_name = 'recibo_crear_item_etnico.html'
    success_url = '/recibo_items_create_etnico/'

    def get_initial(self):
        initial = super().get_initial()
        bodega_predeterminada = Bodega.objects.get(nombre='Ingreso: Compras Efectivas', exportador__nombre='Etnico')
        initial['bodega'] = bodega_predeterminada
        return initial

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Filtrar el queryset para mostrar solo las referencias de la exportadora 'Etnico'
        exportadora_etnico = 'Etnico'
        form.fields['numero_item'].queryset = Referencias.objects.filter(exportador__nombre=exportadora_etnico)
        form.fields['bodega'].queryset = Bodega.objects.filter(exportador__nombre=exportadora_etnico)
        return form

    @transaction.atomic
    def form_valid(self, form):
        numero_item = form.cleaned_data['numero_item']
        form.instance.user = self.request.user
        item = form.save()
        Movimiento.objects.create(
            item_historico=item.numero_item,
            cantidad_cajas_h=item.cantidad_cajas,
            bodega=item.bodega,
            propiedad=item.propiedad,
            fecha_movimiento=item.fecha_movimiento,
            observaciones=item.observaciones,
            fecha=timezone.now(),
            user=item.user
        )
        messages.success(self.request, f'El item {numero_item} se ha creado exitosamente.')
        return JsonResponse({'success': True})

    def form_invalid(self, form):
        return JsonResponse({'success': False, 'html': render_to_string(self.template_name, {'form': form})})


# ----------------------------/// Editar Item Recibo  O de ingreso/// ------------------------------------------------

@method_decorator(login_required, name='dispatch')
class ItemUpdateView(UpdateView):
    model = Item
    form_class = EditarItemForm
    template_name = 'recibo_editar_item.html'
    success_url = '/update_items/'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.object = None

    def get_object(self, queryset=None):
        item_id = int(self.request.POST.get('item_id').replace(".", ""))
        item = get_object_or_404(Item, id=item_id)
        return item

    def get(self, request, *args, **kwargs):
        item_id = int(request.GET.get('item_id').replace(".", ""))
        self.object = get_object_or_404(Item, id=item_id)
        formatted_fecha_movimiento = self.object.fecha_movimiento.strftime('%Y-%m-%d')
        form = self.form_class(
            instance=self.object,
            initial={'fecha_movimiento': formatted_fecha_movimiento}
        )
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return super().get(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['instance'] = self.object
        return kwargs

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        form.instance.user = self.request.user
        numero_item = form.cleaned_data['numero_item']
        item = form.save()
        Movimiento.objects.create(
            item_historico=item.numero_item,
            cantidad_cajas_h=item.cantidad_cajas,
            bodega=item.bodega,
            propiedad=item.propiedad,
            fecha_movimiento=item.fecha_movimiento,
            observaciones=item.observaciones,
            fecha=timezone.now(),
            user=item.user
        )
        messages.success(self.request, f'El item {numero_item} se ha editado exitosamente.')
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse(
                {'success': False, 'html': render_to_string(self.template_name, {'form': form}, request=self.request)})
        else:
            return super().form_invalid(form)


# ----------------------------/// Eliminar Item Recibo  O de ingreso/// ------------------------------------------------
@method_decorator(login_required, name='dispatch')
class ItemDeleteView(UpdateView):
    model = Item
    form_class = EliminarItemForm
    template_name = 'recibo_eliminar_item.html'
    success_url = '/recibo_items_delete/'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.object = None

    def get_object(self, queryset=None):
        item_id = int(self.request.POST.get('item_id').replace(".", ""))
        item = get_object_or_404(Item, id=item_id)
        return item

    def get(self, request, *args, **kwargs):
        item_id = int(request.GET.get('item_id').replace(".", ""))
        self.object = get_object_or_404(Item, id=item_id)
        form = self.form_class(instance=self.object)
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        form.instance.user = self.request.user
        numero_item = form.cleaned_data['numero_item']
        item = form.save()
        item.delete()
        Movimiento.objects.create(
            item_historico=item.numero_item,
            cantidad_cajas_h=item.cantidad_cajas,
            bodega=item.bodega,
            propiedad=item.propiedad,
            fecha_movimiento=item.fecha_movimiento,
            observaciones=item.observaciones,
            fecha=timezone.now(),
            user=item.user
        )
        messages.success(self.request, f'El item {numero_item} se ha eliminado exitosamente.')
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse(
                {'success': False, 'html': render_to_string(self.template_name, {'form': form}, request=self.request)})
        else:
            return super().form_invalid(form)


# -------------------------------------- Vistas Para Fieldex: ---------------------------------------------------------
# ------------------- Lista de Items Fieldex  --------------------------------------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Fieldex'), login_url=reverse_lazy('home')), name='dispatch')
class ItemListViewFieldex(SingleTableView):
    model = Item
    table_class = ItemTable
    template_name = 'recibo_items_list_fieldex.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(bodega__exportador__nombre='Fieldex')
        form = self.form_class(self.request.GET)
        if form.is_valid() and form.cleaned_data.get('item_busqueda'):
            item_busqueda = form.cleaned_data.get('item_busqueda')
            queryset = queryset.filter(numero_item__nombre__icontains=item_busqueda)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# ---------------------------- //// Crear Item Fieldex (Ingreso) //////////////////----------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Fieldex'), login_url=reverse_lazy('home')), name='dispatch')
class ItemCreateViewFieldex(CreateView):
    model = Item
    form_class = ItemForm
    template_name = 'recibo_crear_item_fieldex.html'
    success_url = '/recibo_items_create_fieldex/'

    def get_initial(self):
        initial = super().get_initial()
        bodega_predeterminada = Bodega.objects.get(nombre='Ingreso: Compras Efectivas', exportador__nombre='Fieldex')
        initial['bodega'] = bodega_predeterminada
        return initial

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Filtrar el queryset para mostrar solo las referencias de la exportadora 'Fieldex'
        exportadora_etnico = 'Fieldex'
        form.fields['numero_item'].queryset = Referencias.objects.filter(exportador__nombre=exportadora_etnico)
        form.fields['bodega'].queryset = Bodega.objects.filter(exportador__nombre=exportadora_etnico)
        return form

    @transaction.atomic
    def form_valid(self, form):
        numero_item = form.cleaned_data['numero_item']
        form.instance.user = self.request.user
        item = form.save()
        Movimiento.objects.create(
            item_historico=item.numero_item,
            cantidad_cajas_h=item.cantidad_cajas,
            bodega=item.bodega,
            propiedad=item.propiedad,
            fecha_movimiento=item.fecha_movimiento,
            observaciones=item.observaciones,
            fecha=timezone.now(),
            user=item.user
        )
        messages.success(self.request, f'El item {numero_item} se ha creado exitosamente.')
        return JsonResponse({'success': True})

    def form_invalid(self, form):
        return JsonResponse({'success': False, 'html': render_to_string(self.template_name, {'form': form})})


# -------------------------------------- Vistas Para Juan Matas--------------------------------------------------------


# Mostrar Tabla Recibo - Bodega Recibo Juan Matas (Mostrar Items de ingreso)
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Juan_Matas'), login_url=reverse_lazy('home')), name='dispatch')
class ItemListViewJuan(SingleTableView):
    model = Item
    table_class = ItemTable
    template_name = 'recibo_items_list_juan.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(bodega__exportador__nombre='Juan_Matas')
        form = self.form_class(self.request.GET)
        if form.is_valid() and form.cleaned_data.get('item_busqueda'):
            item_busqueda = form.cleaned_data.get('item_busqueda')
            queryset = queryset.filter(numero_item__nombre__icontains=item_busqueda)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# ---------------------------- //// Crear Item Juan Matas (Ingreso) //////////////////----------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Juan_Matas'), login_url=reverse_lazy('home')), name='dispatch')
class ItemCreateViewJuan(CreateView):
    model = Item
    form_class = ItemForm
    template_name = 'recibo_crear_item_juan.html'
    success_url = '/recibo_items_create_juan/'

    def get_initial(self):
        initial = super().get_initial()
        bodega_predeterminada = Bodega.objects.get(nombre='Ingreso: Compras Efectivas', exportador__nombre='Juan_Matas')
        initial['bodega'] = bodega_predeterminada
        return initial

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Filtrar el queryset para mostrar solo las referencias de la exportadora 'Fieldex'
        exportadora_etnico = 'Juan_Matas'
        form.fields['numero_item'].queryset = Referencias.objects.filter(exportador__nombre=exportadora_etnico)
        form.fields['bodega'].queryset = Bodega.objects.filter(exportador__nombre=exportadora_etnico)
        return form

    @transaction.atomic
    def form_valid(self, form):
        numero_item = form.cleaned_data['numero_item']
        form.instance.user = self.request.user
        item = form.save()
        Movimiento.objects.create(
            item_historico=item.numero_item,
            cantidad_cajas_h=item.cantidad_cajas,
            bodega=item.bodega,
            propiedad=item.propiedad,
            fecha_movimiento=item.fecha_movimiento,
            observaciones=item.observaciones,
            fecha=timezone.now(),
            user=item.user
        )
        messages.success(self.request, f'El item {numero_item} se ha creado exitosamente.')
        return JsonResponse({'success': True})

    def form_invalid(self, form):
        return JsonResponse({'success': False, 'html': render_to_string(self.template_name, {'form': form})})


# -----------------------------------// Tabla General Para Historico de Movimientos----------------------------------//

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
# Tabla De Historico De Movimientos. (Movimientos Inventario General)
class MovimientoListView(SingleTableView):
    model = Movimiento
    table_class = MovimientoTable
    template_name = 'historico.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset()
        form = self.form_class(self.request.GET)
        if form.is_valid():
            item_busqueda = form.cleaned_data.get('item_busqueda')
            if item_busqueda:
                queryset = queryset.filter(item_historico__icontains=item_busqueda)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# ----------------------- ///  Lista De Inventarios Por Bodega /// ----------------------------------------------

# ----------------------- ///  Lista De Inventarios Por Etnico /// ----------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Etnico'), login_url=reverse_lazy('home')), name='dispatch')
class InventarioBodegaEtnicoListView(SingleTableView):
    model = Inventario
    table_class = InventarioTable
    template_name = 'inventario_list_bodega_etnico.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(numero_item__exportador__nombre='Etnico')
        form = self.form_class(self.request.GET)
        if form.is_valid() and form.cleaned_data.get('item_busqueda'):
            item_busqueda = form.cleaned_data.get('item_busqueda')
            queryset = queryset.filter(numero_item__nombre__icontains=item_busqueda)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# ----------------------- ///  Lista De Inventarios Por Fieldex /// ----------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Fieldex'), login_url=reverse_lazy('home')), name='dispatch')
class InventarioBodegaFieldexListView(SingleTableView):
    model = Inventario
    table_class = InventarioTable
    template_name = 'inventario_list_bodega_fieldex.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(numero_item__exportador__nombre='Fieldex')
        form = self.form_class(self.request.GET)
        if form.is_valid() and form.cleaned_data.get('item_busqueda'):
            item_busqueda = form.cleaned_data.get('item_busqueda')
            queryset = queryset.filter(numero_item__nombre__icontains=item_busqueda)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# ----------------------- ///  Lista De Inventarios Por Juan Matas /// ----------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Juan_Matas'), login_url=reverse_lazy('home')), name='dispatch')
class InventarioBodegaJuanListView(SingleTableView):
    model = Inventario
    table_class = InventarioTable
    template_name = 'inventario_list_bodega_juan.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(numero_item__exportador__nombre='Juan_Matas')
        form = self.form_class(self.request.GET)
        if form.is_valid() and form.cleaned_data.get('item_busqueda'):
            item_busqueda = form.cleaned_data.get('item_busqueda')
            queryset = queryset.filter(numero_item__nombre__icontains=item_busqueda)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context



