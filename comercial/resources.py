import locale
import logging
from datetime import timedelta
from decimal import Decimal, InvalidOperation
import openpyxl
from django.core.exceptions import ObjectDoesNotExist
from django.db import models
from django.db.models import Sum, Q
from import_export import resources
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .models import Pedido, Cliente, Fruta, Contenedor, DetallePedido, Iata, Presentacion, Referencias, Exportador, \
    TipoCaja, ClientePresentacion, PresentacionReferencia, AutorizacionCancelacion, Aerolinea, AgenciaCarga, \
    Intermediario, SubExportadora

logger = logging.getLogger(__name__)


# ////////////////////////////////////// Exportaciones De Cartera /////////////////////////////////////////////////////
def obtener_datos_con_totales_cliente(fecha_inicial=None, fecha_final=None, cliente=None, intermediario=None,
                                      grupo=None):
    pedidos_query = Pedido.objects.all()

    if fecha_inicial is not None:
        pedidos_query = pedidos_query.filter(fecha_entrega__gte=fecha_inicial)

    if fecha_final is not None:
        pedidos_query = pedidos_query.filter(fecha_entrega__lte=fecha_final)

    if cliente is not None:
        pedidos_query = pedidos_query.filter(cliente=cliente)

    if intermediario is not None:
        pedidos_query = pedidos_query.filter(intermediario=intermediario)

    if grupo and grupo != "Heavens":
        pedidos_query = pedidos_query.filter(exportadora__nombre=grupo)

    # Obtener los pedidos y sus datos
    pedidos = list(pedidos_query.values(
        'id', 'intermediario__nombre', 'cliente__nombre', 'exportadora__nombre', 'numero_factura', 'awb',
        'fecha_entrega', 'dias_de_vencimiento', 'valor_total_factura_usd',
        'valor_pagado_cliente_usd', 'utilidad_bancaria_usd', 'fecha_pago', 'estado_factura',
        'valor_total_nota_credito_usd', 'descuento', 'nota_credito_no', 'dias_cartera'
    ))

    # Calcular la fecha esperada de pago
    for pedido in pedidos:
        if pedido['fecha_entrega'] and pedido['dias_cartera'] is not None:
            pedido['fecha_esperada_pago'] = pedido['fecha_entrega'] + timedelta(days=pedido['dias_cartera'])
        else:
            pedido['fecha_esperada_pago'] = None

    # Calcular los totales por intermediario, cliente y exportadora
    totales_por_intermediario_cliente_exportadora = pedidos_query.values(
        'intermediario__nombre', 'cliente__nombre', 'exportadora__nombre'
    ).annotate(
        total_factura=Sum('valor_total_factura_usd'),
        total_utilidad=Sum('utilidad_bancaria_usd'),
        total_pagado=Sum('valor_pagado_cliente_usd'),
        total_nc=Sum('valor_total_nota_credito_usd'),
        total_descuentos=Sum('descuento')
    )

    return pedidos, list(totales_por_intermediario_cliente_exportadora)


# -------------------------------- Crear Archivo Excel Ingles Cliente Especifico ------------------------------------

def crear_archivo_excel_cliente(pedidos, totales, ruta_archivo):
    # Crear un nuevo workbook y seleccionar la hoja activa
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Cartera Clientes'

    # Definir y escribir los encabezados de columna
    encabezados = ['Order No.', 'Intermediary', 'Customer', 'Exporter', 'Invoice Number', 'AWB',
                   'Delivery Date',
                   'Days Past Due',
                   'Expected Payment Date', 'Invoice Value USD', 'Amount Paid by Customer USD', 'Credit Note',
                   'Total CN',
                   'Discount', 'Banking Profit', 'Customer Payment Date', 'Invoice Status', 'Balance']
    sheet.append(encabezados)

    # Estilo para los encabezados
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')

    # Estilo para las celdas de datos
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Escribir los datos de los pedidos
    for pedido in pedidos:
        saldo = pedido['valor_total_factura_usd'] - pedido['valor_pagado_cliente_usd'] - pedido[
            'utilidad_bancaria_usd'] - pedido['valor_total_nota_credito_usd'] - pedido['descuento']

        fila = [
            pedido['id'],
            pedido['intermediario__nombre'] if pedido['intermediario__nombre'] else '',
            pedido['cliente__nombre'],
            pedido['exportadora__nombre'],
            pedido['numero_factura'],
            pedido['awb'],
            pedido['fecha_entrega'].strftime('%Y-%m-%d') if pedido['fecha_entrega'] else '',
            pedido['dias_de_vencimiento'],
            pedido['fecha_esperada_pago'].strftime('%Y-%m-%d') if pedido['fecha_esperada_pago'] else '',
            pedido['valor_total_factura_usd'],
            pedido['valor_pagado_cliente_usd'],
            pedido['nota_credito_no'],
            pedido['valor_total_nota_credito_usd'],
            pedido['descuento'],
            pedido['utilidad_bancaria_usd'],
            pedido['fecha_pago'].strftime('%Y-%m-%d') if pedido['fecha_pago'] else '',
            pedido['estado_factura'],
            saldo
        ]
        sheet.append(fila)
        # Aplicar formato de moneda a las celdas relevantes
        moneda_columns = [10, 11, 13, 14, 15, 18]  # Índices de columnas a formatear como moneda
        for col_idx in moneda_columns:
            sheet.cell(row=sheet.max_row, column=col_idx).number_format = '"$"#,##0.00'
        # Centrar el contenido
        for cell in sheet[sheet.max_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # Agregar una línea en blanco antes de los totales
    sheet.append([])

    # Escribir los totales
    totales_encabezado = ['', 'Intermediary', 'Customer', 'Exporter', 'Total Invoices', 'Total Paid by Customer',
                          'Total Banking Profits', 'Total Credit Notes', 'Total Discounts', 'Balance']

    sheet.append(totales_encabezado)

    # Estilo para los encabezados de la hoja de totales
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')

    for total in totales:
        fila_total = [
            '',
            total['intermediario__nombre'] if total['intermediario__nombre'] else '',
            total['cliente__nombre'],
            total['exportadora__nombre'],
            total['total_factura'],
            total['total_pagado'],
            total['total_utilidad'],
            total['total_nc'],
            total['total_descuentos'],
            total['total_factura'] - total['total_pagado'] - total['total_utilidad'] - total['total_nc'] - total[
                'total_descuentos']
        ]
        sheet.append(fila_total)
        moneda_columns_totales = [5, 6, 7, 8, 9, 10]  # Índices de columnas a formatear como moneda en totales
        for col_idx in moneda_columns_totales:
            sheet.cell(row=sheet.max_row, column=col_idx).number_format = '"$"#,##0.00'
        for cell in sheet[sheet.max_row]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='9ed1b7', end_color='9ed1b7',
                                    fill_type='solid')  # Color diferente para totales

    # Ajustar el ancho de las columnas para que se acomoden al contenido
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    # Guardar el archivo
    workbook.save(ruta_archivo)

    return ruta_archivo


# /////////////////////////////////////// Formato Para Enviar Cartera Cliente /////////////////////////////////////////

def obtener_datos_con_totales_enviar_cliente(fecha_inicial=None, fecha_final=None, cliente=None, intermediario=None,
                                             grupo=None):
    pedidos_query = Pedido.objects.exclude(Q(estado_factura="Pagada") | Q(estado_factura="Cancelada"))

    if fecha_inicial is not None:
        pedidos_query = pedidos_query.filter(fecha_entrega__gte=fecha_inicial)

    if fecha_final is not None:
        pedidos_query = pedidos_query.filter(fecha_entrega__lte=fecha_final)

    if cliente is not None:
        pedidos_query = pedidos_query.filter(cliente=cliente)

    if intermediario is not None:
        pedidos_query = pedidos_query.filter(intermediario=intermediario)

    if grupo and grupo != "Heavens":
        pedidos_query = pedidos_query.filter(exportadora__nombre=grupo)

    # Obtener los pedidos y sus datos
    pedidos = list(pedidos_query.values(
        'id', 'intermediario__nombre', 'cliente__nombre', 'exportadora__nombre', 'numero_factura', 'awb',
        'fecha_entrega', 'dias_de_vencimiento', 'valor_total_factura_usd',
        'valor_pagado_cliente_usd', 'utilidad_bancaria_usd', 'fecha_pago', 'estado_factura',
        'valor_total_nota_credito_usd', 'descuento', 'nota_credito_no', 'dias_cartera'
    ))

    # Calcular la fecha esperada de pago
    for pedido in pedidos:
        if pedido['fecha_entrega'] and pedido['dias_cartera'] is not None:
            pedido['fecha_esperada_pago'] = pedido['fecha_entrega'] + timedelta(days=pedido['dias_cartera'])
        else:
            pedido['fecha_esperada_pago'] = None

    # Calcular los totales por intermediario, cliente y exportadora
    totales_por_intermediario_cliente_exportadora = pedidos_query.values(
        'intermediario__nombre', 'cliente__nombre', 'exportadora__nombre'
    ).annotate(
        total_factura=Sum('valor_total_factura_usd'),
        total_utilidad=Sum('utilidad_bancaria_usd'),
        total_pagado=Sum('valor_pagado_cliente_usd'),
        total_nc=Sum('valor_total_nota_credito_usd'),
        total_descuentos=Sum('descuento')
    )

    return pedidos, list(totales_por_intermediario_cliente_exportadora)


def crear_archivo_excel_enviar_cliente(pedidos, totales, ruta_archivo):
    # Crear un nuevo workbook y seleccionar la hoja activa
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Account of statement'

    # Definir y escribir los encabezados de columna
    encabezados = ['Intermediary', 'Customer', 'Exporter', 'AWB', 'Invoice Date', 'Invoice Number',
                   'Invoice Amount USD', 'Amount Paid by Customer USD',
                   'Payment Date', 'Overdue Days', 'Credit Note Number', 'Credit Note Amount',
                   'Final Amount']
    sheet.append(encabezados)

    # Estilo para los encabezados
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')

    # Estilo para las celdas de datos
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Escribir los datos de los pedidos
    for pedido in pedidos:
        total_nc = pedido['valor_total_nota_credito_usd'] + pedido['descuento']
        saldo = (pedido['valor_total_factura_usd'] - pedido['utilidad_bancaria_usd'] -
                 pedido['valor_total_nota_credito_usd'] - pedido['descuento'] - pedido['valor_pagado_cliente_usd'])

        fila = [
            pedido['intermediario__nombre'] if pedido['intermediario__nombre'] else '',
            pedido['cliente__nombre'],
            pedido['exportadora__nombre'],
            pedido['awb'],
            pedido['fecha_entrega'].strftime('%Y-%m-%d') if pedido['fecha_entrega'] else '',
            pedido['numero_factura'],
            pedido['valor_total_factura_usd'],
            pedido['valor_pagado_cliente_usd'],
            pedido['fecha_esperada_pago'].strftime('%Y-%m-%d') if pedido['fecha_esperada_pago'] else '',
            pedido['dias_de_vencimiento'],
            pedido['nota_credito_no'],
            total_nc,
            saldo
        ]
        sheet.append(fila)
        # Aplicar formato de moneda a las celdas relevantes
        moneda_columns = [7, 12, 13]  # Índices de columnas a formatear como moneda
        for col_idx in moneda_columns:
            sheet.cell(row=sheet.max_row, column=col_idx).number_format = '"$"#,##0.00'
        # Centrar el contenido
        for cell in sheet[sheet.max_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # Agregar una línea en blanco antes de los totales
    sheet.append([])

    # Escribir los totales
    totales_encabezado = ['', 'Intermediary', 'Customer', 'Exporter', 'Total Invoices', 'Total Paid by Customer',
                          'Total Banking Profits', 'Total Credit Notes', 'Total Discounts', 'Balance']

    sheet.append(totales_encabezado)

    # Estilo para los encabezados de la hoja de totales
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')

    for total in totales:
        fila_total = [
            '',
            total['intermediario__nombre'] if total['intermediario__nombre'] else '',
            total['cliente__nombre'],
            total['exportadora__nombre'],
            total['total_factura'],
            total['total_pagado'],
            total['total_utilidad'],
            total['total_nc'],
            total['total_descuentos'],
            total['total_factura'] - total['total_pagado'] - total['total_utilidad'] - total['total_nc'] - total[
                'total_descuentos']
        ]
        sheet.append(fila_total)
        moneda_columns_totales = [5, 6, 7, 8, 9, 10]  # Índices de columnas formatear como moneda en totales
        for col_idx in moneda_columns_totales:
            sheet.cell(row=sheet.max_row, column=col_idx).number_format = '"$"#,##0.00'
        for cell in sheet[sheet.max_row]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='9ed1b7', end_color='9ed1b7',
                                    fill_type='solid')  # Color diferente para totales

    # Ajustar el ancho de las columnas para que se acomoden al contenido
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    # Guardar el archivo
    workbook.save(ruta_archivo)

    return ruta_archivo


# ////////////////////////////////////// Fin Exportaciones De Cartera /////////////////////////////////////////////////
# Recursos Para Importaciones.

class ClienteResource(resources.ModelResource):
    class Meta:
        model = Cliente


class ContenedorResource(resources.ModelResource):
    class Meta:
        model = Contenedor


class DetallePedidoResource(resources.ModelResource):
    class Meta:
        model = DetallePedido
        fields = [field.name for field in DetallePedido._meta.fields]

        def before_import_row(self, row, **kwargs):
            """
            Elimina o modifica los valores de los campos no editables antes de importar cada fila,
            y convierte las claves foráneas y campos decimales.
            """
            logger.info(f"Importando fila: {row}")

            # Lista de campos no editables
            campos_no_editables = [field.name for field in DetallePedido._meta.fields if not field.editable]

            for campo in campos_no_editables:
                if campo in row:
                    logger.info(f"Eliminando campo no editable: {campo}")
                    del row[campo]


class ExportadorResource(resources.ModelResource):
    class Meta:
        model = Exportador


class TipoCajaResource(resources.ModelResource):
    class Meta:
        model = TipoCaja


class FrutaResource(resources.ModelResource):
    class Meta:
        model = Fruta


class IataResource(resources.ModelResource):
    class Meta:
        model = Iata


class PedidoResource(resources.ModelResource):
    class Meta:
        model = Pedido
        fields = [field.name for field in Pedido._meta.fields]

    def before_import_row(self, row, **kwargs):
        """
        Elimina o modifica los valores de los campos no editables antes de importar cada fila,
        y convierte las claves foráneas y campos decimales.
        """
        logger.info(f"Importando fila: {row}")

        # Lista de campos no editables
        campos_no_editables = [field.name for field in Pedido._meta.fields if not field.editable]

        for campo in campos_no_editables:
            if campo in row:
                logger.info(f"Eliminando campo no editable: {campo}")
                del row[campo]


class PresentacionResource(resources.ModelResource):
    class Meta:
        model = Presentacion

    def before_import_row(self, row, **kwargs):
        # Convertir campos decimales
        campos_decimales = [field.name for field in Presentacion._meta.fields if isinstance(field, models.DecimalField)]
        for campo in campos_decimales:
            try:
                if row[campo] in [None, '']:
                    row[campo] = None  # Maneja valores vacíos como None
                else:
                    valor = row[campo]
                    # Convertir el valor a float utilizando la configuración regional
                    locale.setlocale(locale.LC_NUMERIC, '')
                    valor = locale.atof(valor)
                    decimal_value = Decimal(valor)
                    # Verificar que el valor no exceda el límite permitido
                    if decimal_value >= Decimal('100000000'):
                        logger.error(f"El valor del campo {campo} excede el límite permitido: {decimal_value}")
                        raise ValueError(f"El valor {decimal_value} del campo {campo} excede el límite permitido")
                    row[campo] = decimal_value
                    logger.info(f"Convertido valor del campo {campo} a Decimal: {row[campo]}")
            except (InvalidOperation, TypeError, ValueError) as e:
                logger.error(f"Error al convertir el campo {campo} a Decimal: {e}")
                raise ValueError(f"El valor del campo {campo} no es válido para Decimal: {row[campo]}")

        super().before_import_row(row, **kwargs)


class ReferenciasResource(resources.ModelResource):
    class Meta:
        model = Referencias

    def before_import_row(self, row, **kwargs):
        # Convertir campos decimales
        campos_decimales = [field.name for field in Referencias._meta.fields if isinstance(field, models.DecimalField)]
        for campo in campos_decimales:
            try:
                if row[campo] in [None, '', "0,00", "0.00"]:
                    row[campo] = Decimal('0')  # Maneja "0,00" y "0.00" específicamente
                else:
                    valor = row[campo]
                    # Convertir el valor a cadena para asegurar el reemplazo adecuado
                    valor = str(valor).replace('.', '').replace(',', '.')
                    # Convertir el valor a float utilizando la configuración regional
                    locale.setlocale(locale.LC_NUMERIC, '')
                    valor = locale.atof(valor)
                    decimal_value = Decimal(valor)
                    # Verificar que el valor no exceda el límite permitido
                    if decimal_value >= Decimal('100000000'):
                        logger.error(f"El valor del campo {campo} excede el límite permitido: {decimal_value}")
                        raise ValueError(f"El valor {decimal_value} del campo {campo} excede el límite permitido")
                    row[campo] = decimal_value
                    logger.info(f"Convertido valor del campo {campo} a Decimal: {row[campo]}")
            except (InvalidOperation, TypeError, ValueError) as e:
                logger.error(f"Error al convertir el campo {campo} a Decimal: {e}")
                raise ValueError(f"El valor del campo {campo} no es válido para Decimal: {row[campo]}")

        super().before_import_row(row, **kwargs)


class ClientePresentacionResource(resources.ModelResource):
    class Meta:
        model = ClientePresentacion


class PresentacionReferenciaResource(resources.ModelResource):
    class Meta:
        model = PresentacionReferencia


class AutorizacionCancelacionResource(resources.ModelResource):
    class Meta:
        model = AutorizacionCancelacion


class AerolineaResource(resources.ModelResource):
    class Meta:
        model = Aerolinea


class AgenciaCargaResource(resources.ModelResource):
    class Meta:
        model = AgenciaCarga


class IntermediarioResource(resources.ModelResource):
    class Meta:
        model = Intermediario


class SubExportadoraResource(resources.ModelResource):
    class Meta:
        model = SubExportadora
