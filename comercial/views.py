import io
import math
import time
from collections import defaultdict
from datetime import datetime
from decimal import Decimal

import pandas as pd
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.decorators import user_passes_test, login_required
from django.db import transaction
from django.db.models import Q, Sum
from django.http import JsonResponse, HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView
from django.views.generic.edit import CreateView, UpdateView
from django_tables2 import SingleTableView
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from xhtml2pdf import pisa
from .forms import SearchForm, PedidoForm, EditarPedidoForm, EliminarPedidoForm, DetallePedidoForm, \
    EliminarDetallePedidoForm, EditarPedidoExportadorForm, EditarDetallePedidoForm, EditarReferenciaForm, \
    EditarPedidoSeguimientoForm, FiltroSemanaExportadoraForm, SearchFormReferencias, EditarPedidoFormDos, \
    EditarPedidoFormCartera, EditarPedidoFormUtilidades, EditarDetallePedidoDosForm, EditarDetallePedidoTresForm, \
    ExportSearchForm, ExportSearchFormSeguimientos
from .models import Pedido, DetallePedido, Referencias, AutorizacionCancelacion, Presentacion, Exportador, Cliente
from .resources import obtener_datos_con_totales_cliente, crear_archivo_excel_cliente, \
    crear_archivo_excel_enviar_cliente, obtener_datos_con_totales_enviar_cliente
from .tables import PedidoTable, DetallePedidoTable, PedidoExportadorTable, CarteraPedidoTable, UtilidadPedidoTable, \
    ResumenPedidoTable, ReferenciasTable, SeguimienosTable, SeguimienosResumenTable


# from .resources import CarteraPedidoResource


# -----------Funcion para permisos por grupo ---------------------
def es_miembro_del_grupo(nombre_grupo):
    def es_miembro(user):
        return user.groups.filter(name=nombre_grupo).exists()

    return es_miembro


@login_required
def redirect_based_on_group_pedidos(request):
    user = request.user
    if user.groups.filter(name='Heavens').exists():
        return redirect('pedido_list_general')
    elif user.groups.filter(name='Fieldex').exists():
        return redirect('pedido_list_fieldex')
    elif user.groups.filter(name='Etnico').exists():
        return redirect('pedido_list_etnico')
    elif user.groups.filter(name='Juan_Matas').exists():
        return redirect('pedido_list_juan')
    else:
        # Redirigir a una vista por defecto si el usuario no pertenece a ninguno de los grupos
        return redirect('home')


@login_required
def redirect_based_on_group_cartera(request):
    user = request.user
    if user.groups.filter(name='Heavens').exists():
        return redirect('cartera_list_heavens')
    elif user.groups.filter(name='Fieldex').exists():
        return redirect('cartera_list_fieldex')
    elif user.groups.filter(name='Etnico').exists():
        return redirect('cartera_list_etnico')
    elif user.groups.filter(name='Juan_Matas').exists():
        return redirect('cartera_list_juan')
    else:
        # Redirigir a una vista por defecto si el usuario no pertenece a ninguno de los grupos
        return redirect('home')


# ----------------- Resumen Exportaciones Table View -------------------------------------

@method_decorator(login_required, name='dispatch')
class ResumenPedidoListView(SingleTableView):
    model = DetallePedido
    table_class = ResumenPedidoTable
    template_name = 'resumen_pedido.html'

    def dispatch(self, request, *args, **kwargs):
        pedido_id = self.kwargs.get('pedido_id')
        pedido = get_object_or_404(Pedido, pk=pedido_id)
        # Comprueba si el usuario pertenece al grupo requerido
        if not request.user.groups.filter(name=pedido.exportadora.nombre).exists():
            return HttpResponseForbidden("No tienes permiso para ver estos detalles del pedido")

        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        pedido_id = self.kwargs.get('pedido_id')
        return DetallePedido.objects.filter(pedido__id=pedido_id)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pedido_id = self.kwargs.get('pedido_id')
        context['pedido'] = get_object_or_404(Pedido, pk=pedido_id)
        queryset = self.get_queryset()

        # Calcular el total de cajas solicitadas
        total_cajas_solicitadas = self.get_queryset().aggregate(Sum('cajas_solicitadas'))['cajas_solicitadas__sum']
        total_peso_bruto = sum(obj.calcular_peso_bruto() for obj in queryset)
        total_piezas = math.ceil(float(sum(obj.calcular_no_piezas() for obj in queryset)))
        context['total_cajas_solicitadas'] = total_cajas_solicitadas
        context['total_peso_bruto'] = total_peso_bruto
        context['total_piezas'] = total_piezas

        return context


# ---------------------------------Resumen Exportaciones PDF -----------------------------------------------------
def exportar_detalle_pedido_a_pdf(request):
    # Crear una respuesta HTTP para un documento PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="detalle_pedido.pdf"'

    # Crear un objeto canvas de ReportLab para dibujar el PDF
    p = canvas.Canvas(response, pagesize=letter)

    # Inicializar la posición vertical para dibujar en la página
    y = 750

    # Encabezados de las columnas
    encabezados = [
        "Fruta", "Presentación", "Cajas Solicitadas", "Peso Caja", "Kilos", "Cajas Enviadas",
        "Marca Caja", "Referencia", "Stickers", "Lleva Contenedor", "Referencia Contenedor",
        "$Utilidad Por Caja", "$Por Caja USD", "$Por Producto"
    ]

    # Dibujar los encabezados de las columnas
    for encabezado in encabezados:
        p.drawString(30, y, encabezado)
        y -= 20

    # Resetear la posición Y para los datos
    y = 700

    # Obtener y dibujar los datos
    for detalle in DetallePedido.objects.filter(pedido_id=166):
        datos = [
            str(detalle.fruta), str(detalle.presentacion), detalle.cajas_solicitadas,
            detalle.presentacion_peso, detalle.kilos, detalle.cajas_enviadas,
            str(detalle.tipo_caja), str(detalle.referencia), detalle.stickers,
            "Sí" if detalle.lleva_contenedor else "No", detalle.referencia_contenedor,
            detalle.tarifa_utilidad, detalle.valor_x_caja_usd, detalle.valor_x_producto
        ]

        for dato in datos:
            p.drawString(30, y, str(dato))
            y -= 20
            if y < 50:  # Cambiar de página si es necesario
                p.showPage()
                y = 750

    # Finalizar el PDF
    p.showPage()
    p.save()

    # Devolver la respuesta
    return response


# ------------------ Exportacion de Utilidades Excel General --------------------------------------------------------
class ExportarUtilidadesView(TemplateView):
    template_name = 'export_utilidades_general.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Agrega contexto adicional aquí si es necesario
        return context


# ---------------------------------- Funcion que exporta las utilidades a Excel ---------------------------------------

@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Heavens'), login_url='home'))
def exportar_utilidades_excel(request):
    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()

    # Hoja 1: Utilidades Totales General
    worksheet1 = workbook.active
    worksheet1.title = 'Utilidades Totales General'
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    total_align = Alignment(horizontal="center")

    # Definir el color de relleno rojo suave
    fill_red_soft = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Encabezados
    columns = ['No. Pedido', 'Fecha Entrega Pedido', 'Cliente', 'Exportador', 'AWB', 'Fecha Pago Cliente', 'No Factura',
               'Valor Total Factura USD', 'Valor Pagado Cliente', 'Estado Factura', 'T Cajas Enviadas',
               'Trm Monetizacion',
               'TRM Banrep', 'Valor Utilidad USD', 'Valor Utilidad Pesos', 'Documento Cobro Utilidad',
               'Fecha Pago Utilidad', 'Diferencia O Abono', 'Estado Utilidad', 'Cobrar Utilidad']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet1.cell(row=1, column=col_num, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Crear un diccionario para almacenar los totales de Utilidades por exportadora
    totales_por_utilidad_usd = defaultdict(Decimal)
    totales_no_cobrables_por_exportadora = defaultdict(Decimal)
    totales_cobrados_por_exportadora = defaultdict(Decimal)
    totales_por_cobrar_por_exportadora = defaultdict(Decimal)

    fecha_inicial_str = request.POST.get('fecha_inicial')
    fecha_final_str = request.POST.get('fecha_final')

    # Verificar si las fechas están vacías
    if fecha_inicial_str and fecha_final_str:
        # Convertir las cadenas de fecha en datetime
        fecha_inicial = datetime.strptime(fecha_inicial_str, '%Y-%m-%d')
        fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d')

        # Filtrar los pedidos por fecha_entrega dentro del rango
        queryset = Pedido.objects.filter(fecha_entrega__gte=fecha_inicial, fecha_entrega__lte=fecha_final)
    else:
        # Si las fechas están vacías, exportar todos los pedidos
        queryset = Pedido.objects.all()

    # Obtener los datos
    for pedido in queryset:
        valor_utilidad_usd = pedido.valor_total_utilidad_usd
        if valor_utilidad_usd is None:
            continue  # O puedes manejarlo de alguna otra manera según tu lógica de negocio
        if pedido.estado_utilidad == "Factura en abono" or pedido.estado_utilidad == "Pendiente Pago Cliente":
            totales_no_cobrables_por_exportadora[pedido.exportadora.nombre] += Decimal(valor_utilidad_usd)
        if pedido.fecha_pago_utilidad is not None and pedido.documento_cobro_utilidad is not None:
            totales_cobrados_por_exportadora[pedido.exportadora.nombre] += Decimal(valor_utilidad_usd)
        if pedido.fecha_pago_utilidad is None and pedido.estado_factura == "Pagada" and (
                pedido.estado_utilidad == "Por Facturar" or pedido.estado_utilidad == "Facturada"):
            totales_por_cobrar_por_exportadora[pedido.exportadora.nombre] += Decimal(valor_utilidad_usd)
        totales_por_utilidad_usd[pedido.exportadora.nombre] += Decimal(valor_utilidad_usd)

    # Agregar datos a la hoja de trabajo 1
    for row_num, pedido in enumerate(queryset, start=2):
        cobrar_utilidad = "Sí" if pedido.estado_utilidad == "Por Facturar" or pedido.estado_utilidad == "Facturada" else "No"
        row = [
            pedido.pk,
            pedido.fecha_entrega,
            pedido.cliente.nombre,
            pedido.exportadora.nombre,
            pedido.awb,
            pedido.fecha_pago,
            pedido.numero_factura,
            pedido.valor_total_factura_usd,
            pedido.valor_pagado_cliente_usd,
            pedido.estado_factura,
            pedido.total_cajas_enviadas,
            pedido.trm_monetizacion,
            pedido.tasa_representativa_usd_diaria,
            pedido.valor_total_utilidad_usd,
            pedido.valor_utilidad_pesos,
            pedido.documento_cobro_utilidad,
            pedido.fecha_pago_utilidad,
            pedido.diferencia_por_abono,
            pedido.estado_utilidad,
            cobrar_utilidad,  # Añadido valor de 'Cobrar utilidad'
        ]
        for col_num, cell_value in enumerate(row, start=1):
            cell = worksheet1.cell(row=row_num, column=col_num, value=cell_value)
            # Aplicar formato de moneda a las columnas específicas
            if col_num in [8, 9, 11, 12, 13, 14, 18]:
                cell.number_format = '$#,##0.00'
            # Pintar la fila si el numero_factura es 'Pedido Cancelado'
            if row[6] == 'Pedido Cancelado':  # '6' es el índice de 'No Factura'
                cell.fill = fill_red_soft

    # Hoja 2: Totales por Exportadora
    worksheet2 = workbook.create_sheet(title='Totales por Exportadora')

    # Encabezados para la segunda hoja
    totals_columns = ["Exportadora", "Total Utilidades USD", "Total Utilidades No Cobrables USD",
                      "Total Utilidades Cobradas USD", "Total Por Cobrar USD"]
    for col_num, column_title in enumerate(totals_columns, start=1):
        cell = worksheet2.cell(row=1, column=col_num, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Función para dar estilo a los totales.
    def aplicar_estilo_total(worksheet, fila):
        for col in range(1, len(totals_columns) + 1):
            celda = worksheet.cell(row=fila, column=col)
            celda.font = total_font
            celda.fill = total_fill
            celda.alignment = total_align

    # Agregar totales a la segunda hoja
    row_num = 2
    for exportadora in totales_por_utilidad_usd.keys():
        worksheet2.cell(row=row_num, column=1, value=exportadora)
        for col_num, total in enumerate([totales_por_utilidad_usd[exportadora],
                                         totales_no_cobrables_por_exportadora[exportadora],
                                         totales_cobrados_por_exportadora[exportadora],
                                         totales_por_cobrar_por_exportadora[exportadora]], start=2):
            cell = worksheet2.cell(row=row_num, column=col_num, value=total)
            cell.number_format = '$#,##0.00'
        aplicar_estilo_total(worksheet2, row_num)
        row_num += 1

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="utilidades_pedidos_general.xlsx"'

    return response


# ------------------ Exportacion de utilidades Excel Etnico --------------------------------------------------------

class ExportarUtilidadesEtnicoView(TemplateView):
    template_name = 'export_utilidades_etnico.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Agrega contexto adicional aquí si es necesario
        return context


@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Etnico'), login_url='home'))
def exportar_utilidades_etnico(request):
    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Utilidades Totales Etnico'
    font = Font(bold=True)
    fill = PatternFill(start_color="3ef983", end_color="3ef983", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="3580e0", end_color="3580e0", fill_type="solid")
    total_align = Alignment(horizontal="center")
    # Definir el color de relleno rojo suave
    fill_red_soft = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Encabezados
    columns = ['No. Pedido', 'Fecha Entrega Pedido', 'Cliente', 'Exportador', 'AWB', 'Fecha Pago Cliente', 'No Factura',
               'Valor Total Factura USD', 'Valor Pagado Cliente', 'Estado Factura', 'T Cajas Enviadas',
               'Trm Monetizacion',
               'TRM Banrep', 'Valor Utilidad USD', 'Valor Utilidad Pesos', 'Documento Cobro Utilidad',
               'Fecha Pago Utilidad', 'Diferencia O Abono', 'Estado Utilidad', 'Cobrar Utilidad']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill
    # Crear un diccionario para almacenar los totales de Utilidades por exportadora
    totales_por_utilidad_usd = defaultdict(Decimal)
    totales_no_cobrables_por_exportadora = defaultdict(Decimal)
    totales_cobrados_por_exportadora = defaultdict(Decimal)
    totales_por_cobrar_por_exportadora = defaultdict(Decimal)

    fecha_inicial_str = request.POST.get('fecha_inicial')
    fecha_final_str = request.POST.get('fecha_final')

    # Verificar si las fechas están vacías
    if fecha_inicial_str and fecha_final_str:
        # Convertir las cadenas de fecha en datetime
        fecha_inicial = datetime.strptime(fecha_inicial_str, '%Y-%m-%d')
        fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d')

        # Filtrar los pedidos por fecha_entrega dentro del rango
        queryset = Pedido.objects.filter(
            Q(fecha_entrega__gte=fecha_inicial, fecha_entrega__lte=fecha_final) | Q(exportadora__nombre='Etnico'))
    else:
        # Si las fechas están vacías, exportar todos los pedidos
        queryset = Pedido.objects.filter(exportadora__nombre='Etnico')

        # Obtener los datos de tu modelo y calcular los totales
        for pedido in queryset:
            valor_utilidad_usd = pedido.valor_total_utilidad_usd
            if valor_utilidad_usd is None:
                continue  # O puedes manejarlo de alguna otra manera según tu lógica de negocio
            if pedido.estado_utilidad == "Factura en abono" or pedido.estado_utilidad == "Pendiente Pago Cliente":
                totales_no_cobrables_por_exportadora[pedido.exportadora.nombre] += Decimal(valor_utilidad_usd)
            if pedido.fecha_pago_utilidad is not None and pedido.documento_cobro_utilidad is not None:
                totales_cobrados_por_exportadora[pedido.exportadora.nombre] += Decimal(valor_utilidad_usd)
            if pedido.fecha_pago_utilidad is None and pedido.estado_factura == "Pagada" and (
                    pedido.estado_utilidad == "Por Facturar" or pedido.estado_utilidad == "Facturada"):
                totales_por_cobrar_por_exportadora[pedido.exportadora.nombre] += Decimal(valor_utilidad_usd)
            totales_por_utilidad_usd[pedido.exportadora.nombre] += Decimal(valor_utilidad_usd)

        # Agregar datos a la hoja de trabajo 1
        for row_num, pedido in enumerate(queryset, start=2):
            cobrar_utilidad = "Sí" if pedido.estado_utilidad == "Por Facturar" or pedido.estado_utilidad == "Facturada" else "No"
            row = [
                pedido.pk,
                pedido.fecha_entrega,
                pedido.cliente.nombre,
                pedido.exportadora.nombre,
                pedido.awb,
                pedido.fecha_pago,
                pedido.numero_factura,
                pedido.valor_total_factura_usd,
                pedido.valor_pagado_cliente_usd,
                pedido.estado_factura,
                pedido.total_cajas_enviadas,
                pedido.trm_monetizacion,
                pedido.tasa_representativa_usd_diaria,
                pedido.valor_total_utilidad_usd,
                pedido.valor_utilidad_pesos,
                pedido.documento_cobro_utilidad,
                pedido.fecha_pago_utilidad,
                pedido.diferencia_por_abono,
                pedido.estado_utilidad,
                cobrar_utilidad,  # Añadido valor de 'Cobrar utilidad'
            ]
            for col_num, cell_value in enumerate(row, start=1):
                cell = worksheet.cell(row=row_num, column=col_num, value=cell_value)
                # Aplicar formato de moneda a las columnas específicas
                if col_num in [8, 9, 11, 12, 13, 14, 18]:
                    cell.number_format = '$#,##0.00'
                # Pintar la fila si el numero_factura es 'Pedido Cancelado'
                if row[6] == 'Pedido Cancelado':  # '6' es el índice de 'No Factura'
                    cell.fill = fill_red_soft

    def aplicar_estilo_total(fila):  # Funcion APara Dar Estilo A Los Totales.
        for col in range(1, len(columns) + 1):
            celda = worksheet.cell(row=fila, column=col)
            celda.font = total_font
            celda.fill = total_fill
            celda.alignment = total_align

    row_num += 2  # Saltar a la siguiente fila después de los datos
    for exportadora, total in totales_por_utilidad_usd.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Utilidades USD")
        worksheet.cell(row=row_num, column=3, value=total)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila
    for exportadora, total_no_cobrable in totales_no_cobrables_por_exportadora.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Utilidades No Cobrables USD")
        worksheet.cell(row=row_num, column=3, value=total_no_cobrable)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila
    for exportadora, total_cobrado in totales_cobrados_por_exportadora.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Utilidades Pagadas")
        worksheet.cell(row=row_num, column=3, value=total_cobrado)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila
    for exportadora, total_por_cobrar in totales_por_cobrar_por_exportadora.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Utilidades Por Pagar")
        worksheet.cell(row=row_num, column=3, value=total_por_cobrar)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="utilidades_pedidos_etnico.xlsx"'

    return response


# ------------------ Exportacion de Utilidades Excel Fieldex --------------------------------------------------------

class ExportarUtilidadesFieldexView(TemplateView):
    template_name = 'export_utilidades_fieldex.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Agrega contexto adicional aquí si es necesario
        return context


@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Fieldex'), login_url='home'))
def exportar_utilidades_fieldex(request):
    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Utilidades Totales Fieldex'
    font = Font(bold=True)
    fill = PatternFill(start_color="3ef983", end_color="3ef983", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="3580e0", end_color="3580e0", fill_type="solid")
    total_align = Alignment(horizontal="center")
    # Definir el color de relleno rojo suave
    fill_red_soft = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Encabezados
    columns = ['No. Pedido', 'Fecha Entrega Pedido', 'Cliente', 'Exportador', 'AWB', 'Fecha Pago Cliente', 'No Factura',
               'Valor Total Factura USD', 'Valor Pagado Cliente', 'Estado Factura', 'T Cajas Enviadas',
               'Trm Monetizacion',
               'TRM Banrep', 'Valor Utilidad USD', 'Valor Utilidad Pesos', 'Documento Cobro Utilidad',
               'Fecha Pago Utilidad', 'Diferencia O Abono', 'Estado Utilidad', 'Cobrar Utilidad']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill
    # Crear un diccionario para almacenar los totales de Utilidades por exportadora
    totales_por_utilidad_usd = defaultdict(Decimal)
    totales_no_cobrables_por_exportadora = defaultdict(Decimal)
    totales_cobrados_por_exportadora = defaultdict(Decimal)
    totales_por_cobrar_por_exportadora = defaultdict(Decimal)
    fecha_inicial_str = request.POST.get('fecha_inicial')
    fecha_final_str = request.POST.get('fecha_final')

    # Verificar si las fechas están vacías
    if fecha_inicial_str and fecha_final_str:
        # Convertir las cadenas de fecha en datetime
        fecha_inicial = datetime.strptime(fecha_inicial_str, '%Y-%m-%d')
        fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d')

        # Filtrar los pedidos por fecha_entrega dentro del rango
        queryset = Pedido.objects.filter(
            Q(fecha_entrega__gte=fecha_inicial, fecha_entrega__lte=fecha_final) | Q(exportadora__nombre='Fieldex'))
    else:
        # Si las fechas están vacías, exportar todos los pedidos
        queryset = Pedido.objects.filter(exportadora__nombre='Fieldex')

    # Obtener los datos de tu modelo y calcular los totales
    for pedido in queryset:
        valor_utilidad_usd = pedido.valor_total_utilidad_usd
        if valor_utilidad_usd is None:
            continue  # O puedes manejarlo de alguna otra manera según tu lógica de negocio
        if pedido.estado_utilidad == "Factura en abono" or pedido.estado_utilidad == "Pendiente Pago Cliente":
            totales_no_cobrables_por_exportadora[pedido.exportadora.nombre] += Decimal(valor_utilidad_usd)
        if pedido.fecha_pago_utilidad is not None and pedido.documento_cobro_utilidad is not None:
            totales_cobrados_por_exportadora[pedido.exportadora.nombre] += Decimal(valor_utilidad_usd)
        if pedido.fecha_pago_utilidad is None and pedido.estado_factura == "Pagada" and (
                pedido.estado_utilidad == "Por Facturar" or pedido.estado_utilidad == "Facturada"):
            totales_por_cobrar_por_exportadora[pedido.exportadora.nombre] += Decimal(valor_utilidad_usd)
        totales_por_utilidad_usd[pedido.exportadora.nombre] += Decimal(valor_utilidad_usd)

    # Agregar datos a la hoja de trabajo 1
    for row_num, pedido in enumerate(queryset, start=2):
        cobrar_utilidad = "Sí" if pedido.estado_utilidad == "Por Facturar" or pedido.estado_utilidad == "Facturada" else "No"
        row = [
            pedido.pk,
            pedido.fecha_entrega,
            pedido.cliente.nombre,
            pedido.exportadora.nombre,
            pedido.awb,
            pedido.fecha_pago,
            pedido.numero_factura,
            pedido.valor_total_factura_usd,
            pedido.valor_pagado_cliente_usd,
            pedido.estado_factura,
            pedido.total_cajas_enviadas,
            pedido.trm_monetizacion,
            pedido.tasa_representativa_usd_diaria,
            pedido.valor_total_utilidad_usd,
            pedido.valor_utilidad_pesos,
            pedido.documento_cobro_utilidad,
            pedido.fecha_pago_utilidad,
            pedido.diferencia_por_abono,
            pedido.estado_utilidad,
            cobrar_utilidad,  # Añadido valor de 'Cobrar utilidad'
        ]
        for col_num, cell_value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_num, column=col_num, value=cell_value)
            # Aplicar formato de moneda a las columnas específicas
            if col_num in [8, 9, 11, 12, 13, 14, 18]:
                cell.number_format = '$#,##0.00'
            # Pintar la fila si el numero_factura es 'Pedido Cancelado'
            if row[6] == 'Pedido Cancelado':  # '6' es el índice de 'No Factura'
                cell.fill = fill_red_soft

    # Agregar los totales al final de la hoja de trabajo

    def aplicar_estilo_total(fila):  # Funcion APara Dar Estilo A Los Totales.
        for col in range(1, len(columns) + 1):
            celda = worksheet.cell(row=fila, column=col)
            celda.font = total_font
            celda.fill = total_fill
            celda.alignment = total_align

    row_num += 2  # Saltar a la siguiente fila después de los datos
    for exportadora, total in totales_por_utilidad_usd.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Utilidades USD")
        worksheet.cell(row=row_num, column=3, value=total)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila
    for exportadora, total_no_cobrable in totales_no_cobrables_por_exportadora.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Utilidades No Cobrables USD")
        worksheet.cell(row=row_num, column=3, value=total_no_cobrable)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila
    for exportadora, total_cobrado in totales_cobrados_por_exportadora.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Utilidades Pagadas")
        worksheet.cell(row=row_num, column=3, value=total_cobrado)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila
    for exportadora, total_por_cobrar in totales_por_cobrar_por_exportadora.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Utilidades Por Pagar")
        worksheet.cell(row=row_num, column=3, value=total_por_cobrar)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="utilidades_pedidos_fieldex.xlsx"'

    return response


# ------------------ Exportacion de Utilidades Excel Juan Matas --------------------------------------------------------
class ExportarUtilidadesJuanView(TemplateView):
    template_name = 'export_utilidades_juan.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Agrega contexto adicional aquí si es necesario
        return context


@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Juan_Matas'), login_url='home'))
def exportar_utilidades_juan(request):
    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Utilidades Totales Juan Matas'
    font = Font(bold=True)
    fill = PatternFill(start_color="3ef983", end_color="3ef983", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="3580e0", end_color="3580e0", fill_type="solid")
    total_align = Alignment(horizontal="center")
    # Definir el color de relleno rojo suave
    fill_red_soft = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Encabezados
    columns = ['No. Pedido', 'Fecha Entrega Pedido', 'Cliente', 'Exportador', 'AWB', 'Fecha Pago Cliente', 'No Factura',
               'Valor Total Factura USD', 'Valor Pagado Cliente', 'Estado Factura', 'T Cajas Enviadas',
               'Trm Monetizacion',
               'TRM Banrep', 'Valor Utilidad USD', 'Valor Utilidad Pesos', 'Documento Cobro Utilidad',
               'Fecha Pago Utilidad', 'Diferencia O Abono', 'Estado Utilidad', 'Cobrar Utilidad']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill
    # Crear un diccionario para almacenar los totales de Utilidades por exportadora
    totales_por_utilidad_usd = defaultdict(Decimal)
    totales_no_cobrables_por_exportadora = defaultdict(Decimal)
    totales_cobrados_por_exportadora = defaultdict(Decimal)
    totales_por_cobrar_por_exportadora = defaultdict(Decimal)
    fecha_inicial_str = request.POST.get('fecha_inicial')
    fecha_final_str = request.POST.get('fecha_final')

    # Verificar si las fechas están vacías
    if fecha_inicial_str and fecha_final_str:
        # Convertir las cadenas de fecha en datetime
        fecha_inicial = datetime.strptime(fecha_inicial_str, '%Y-%m-%d')
        fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d')

        # Filtrar los pedidos por fecha_entrega dentro del rango
        queryset = Pedido.objects.filter(
            Q(fecha_entrega__gte=fecha_inicial, fecha_entrega__lte=fecha_final) | Q(exportadora__nombre='Juan_Matas'))
    else:
        # Si las fechas están vacías, exportar todos los pedidos
        queryset = Pedido.objects.filter(exportadora__nombre='Juan_Matas')

    # Obtener los datos de tu modelo y calcular los totales
    for pedido in queryset:
        valor_utilidad_usd = pedido.valor_total_utilidad_usd
        if valor_utilidad_usd is None:
            continue  # O puedes manejarlo de alguna otra manera según tu lógica de negocio
        if pedido.estado_utilidad == "Factura en abono" or pedido.estado_utilidad == "Pendiente Pago Cliente":
            totales_no_cobrables_por_exportadora[pedido.exportadora.nombre] += Decimal(valor_utilidad_usd)
        if pedido.fecha_pago_utilidad is not None and pedido.documento_cobro_utilidad is not None:
            totales_cobrados_por_exportadora[pedido.exportadora.nombre] += Decimal(valor_utilidad_usd)
        if pedido.fecha_pago_utilidad is None and pedido.estado_factura == "Pagada" and (
                pedido.estado_utilidad == "Por Facturar" or pedido.estado_utilidad == "Facturada"):
            totales_por_cobrar_por_exportadora[pedido.exportadora.nombre] += Decimal(valor_utilidad_usd)
        totales_por_utilidad_usd[pedido.exportadora.nombre] += Decimal(valor_utilidad_usd)

    # Agregar datos a la hoja de trabajo 1
    for row_num, pedido in enumerate(queryset, start=2):
        cobrar_utilidad = "Sí" if pedido.estado_utilidad == "Por Facturar" or pedido.estado_utilidad == "Facturada" else "No"
        row = [
            pedido.pk,
            pedido.fecha_entrega,
            pedido.cliente.nombre,
            pedido.exportadora.nombre,
            pedido.awb,
            pedido.fecha_pago,
            pedido.numero_factura,
            pedido.valor_total_factura_usd,
            pedido.valor_pagado_cliente_usd,
            pedido.estado_factura,
            pedido.total_cajas_enviadas,
            pedido.trm_monetizacion,
            pedido.tasa_representativa_usd_diaria,
            pedido.valor_total_utilidad_usd,
            pedido.valor_utilidad_pesos,
            pedido.documento_cobro_utilidad,
            pedido.fecha_pago_utilidad,
            pedido.diferencia_por_abono,
            pedido.estado_utilidad,
            cobrar_utilidad,  # Añadido valor de 'Cobrar utilidad'
        ]
        for col_num, cell_value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_num, column=col_num, value=cell_value)
            # Aplicar formato de moneda a las columnas específicas
            if col_num in [8, 9, 11, 12, 13, 14, 18]:
                cell.number_format = '$#,##0.00'
            # Pintar la fila si el numero_factura es 'Pedido Cancelado'
            if row[6] == 'Pedido Cancelado':  # '6' es el índice de 'No Factura'
                cell.fill = fill_red_soft

    # Agregar los totales al final de la hoja de trabajo

    def aplicar_estilo_total(fila):  # Funcion APara Dar Estilo A Los Totales.
        for col in range(1, len(columns) + 1):
            celda = worksheet.cell(row=fila, column=col)
            celda.font = total_font
            celda.fill = total_fill
            celda.alignment = total_align

    row_num += 2  # Saltar a la siguiente fila después de los datos
    for exportadora, total in totales_por_utilidad_usd.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Utilidades USD")
        worksheet.cell(row=row_num, column=3, value=total)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila
    for exportadora, total_no_cobrable in totales_no_cobrables_por_exportadora.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Utilidades No Cobrables USD")
        worksheet.cell(row=row_num, column=3, value=total_no_cobrable)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila
    for exportadora, total_cobrado in totales_cobrados_por_exportadora.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Utilidades Pagadas")
        worksheet.cell(row=row_num, column=3, value=total_cobrado)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila
    for exportadora, total_por_cobrar in totales_por_cobrar_por_exportadora.items():
        worksheet.cell(row=row_num, column=1, value=exportadora)
        worksheet.cell(row=row_num, column=2, value="Total Utilidades Por Pagar")
        worksheet.cell(row=row_num, column=3, value=total_por_cobrar)
        aplicar_estilo_total(row_num)
        row_num += 1  # Prepararse para la siguiente fila

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="utilidades_pedidos_juan.xlsx"'

    return response


# ----------------------- Vista exportacion Detalles de pedido --------------------------------------------
class ExportarDetallesPedidoView(TemplateView):
    template_name = 'export_detalles_pedido_view.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Agrega contexto adicional aquí si es necesario
        return context


# ------------------ Exportacion de Detalles de Pedidos Excel General -------------------------------------
@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Heavens'), login_url='home'))
def exportar_detalles_pedidos_excel(request):
    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Detalles De Pedidos General'
    font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="251819", end_color="251819", fill_type="solid")

    # Encabezados
    columns = ['Pedido', 'F Entrega', 'Exportador', 'Cliente', 'Fruta', 'Presentacion', 'Cajas Solicitadas',
               'Peso Presentacion', 'kilos', 'Cajas Enviadas', 'Kilos Enviados', 'Diferencia', 'Tipo Caja',
               'Referencia', 'Stiker', 'Lleva Contenedor', 'Ref Contenedor', 'Cant Contenedor', 'Tarifa utilidad',
               'Valor x Caja USD', 'Valor X Producto', 'No Cajas NC', 'Valor NC', 'Afecta utilidad',
               'Valor Total utilidad Producto', 'Precio Proforma', 'Observaciones']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill

    # Filtrar datos basado en el rango de número de pedido
    numero_pedido_inicial = request.POST.get('numero_pedido_inicial')
    numero_pedido_final = request.POST.get('numero_pedido_final')

    queryset = DetallePedido.objects.all()

    if numero_pedido_inicial and numero_pedido_final:
        queryset = queryset.filter(pedido__pk__range=(numero_pedido_inicial, numero_pedido_final))

    # Agregar datos al libro de trabajo
    for row_num, detalle in enumerate(queryset, start=2):
        row = [
            detalle.pedido.pk,
            detalle.pedido.fecha_entrega,
            detalle.pedido.exportadora.nombre,
            detalle.pedido.cliente.nombre,
            detalle.fruta.nombre,
            detalle.presentacion.nombre,
            detalle.cajas_solicitadas,
            detalle.presentacion_peso,
            detalle.kilos,
            detalle.cajas_enviadas,
            detalle.kilos_enviados,
            detalle.diferencia,
            detalle.tipo_caja.nombre,
            detalle.referencia.nombre,
            detalle.stickers,
            detalle.lleva_contenedor,
            detalle.referencia_contenedor,
            detalle.cantidad_contenedores,
            detalle.tarifa_utilidad,
            detalle.valor_x_caja_usd,
            detalle.valor_x_producto,
            detalle.no_cajas_nc,
            detalle.valor_nota_credito_usd,
            detalle.afecta_utilidad,
            detalle.valor_total_utilidad_x_producto,
            detalle.precio_proforma,
            detalle.observaciones,
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Detalles_Pedidos.xlsx"'

    return response


# ------------------ Exportacion de Pedidos Excel General --------------------------------------------------------

# Vista Para Exportar Pedidos:

class ExportarPedidosView(TemplateView):
    template_name = 'export_pedidos_general.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Agrega contexto adicional aquí si es necesario
        return context


# Exportacion de Pedidos OpenPyXl >
@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Heavens'), login_url='home'))
def exportar_pedidos_excel(request):
    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Pedidos Totales General'
    font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="1e0c42", end_color="1e0c42", fill_type="solid")

    # Encabezados
    columns = ['No', 'Cliente', 'Semana',
               'Fecha Solicitud', 'Fecha Entrega', 'Fecha Llegada', 'Exportador', 'Subexportadora', 'Intermediario',
               'Dias Cartera',
               'AWB', 'Destino', 'Aerolinea', 'Agencia De Carga', 'Responsable Reserva', 'Numero Factura',
               'Total Cajas Solicitadas', 'Total Cajas Enviadas', 'Peso Bruto Solicitado', 'Peso Bruto Enviado',
               'Pallets Solicitados', 'Pallets Enviados', 'Peso AWB', 'ETA', 'ETD', 'Variedades', 'Descuento Comercial',
               'No NC', 'Motivo NC', 'Valor Total NC', 'Valor Pagado Cliente', 'Estado Factura',
               'Utilidad Bancaria USD', 'Fecha Pago Cliente', 'TRM Monetización', 'Fecha Monetización', 'Trm Banrep',
               'Trm Cotización', 'Diferencia Pago', 'Dias Vencimiento', 'Valor Total Factura USD', 'Valor Utilidad USD',
               'Valor Utilidad Pesos', 'Documento Cobro Utilidad', 'Fecha Pago Utilidad', 'Estado Utilidad',
               'Estado Cancelacion', 'Estado Documentos', 'Estado Reserva', 'Termo', 'Diferencia AWB/Factura',
               'Eta Real',
               'Estado Pedido', 'Observaciones Tracking',
               'Observaciones Generales']

    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill

    fecha_inicial_str = request.POST.get('fecha_inicial')
    fecha_final_str = request.POST.get('fecha_final')

    # Verificar si las fechas están vacías o nulas
    if fecha_inicial_str and fecha_final_str:
        # Convertir las cadenas de fecha en objetos datetime
        fecha_inicial = datetime.strptime(fecha_inicial_str, '%Y-%m-%d')
        fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d')

        # Filtrar los pedidos por fecha_entrega dentro del rango
        queryset = Pedido.objects.filter(fecha_entrega__gte=fecha_inicial, fecha_entrega__lte=fecha_final)
    else:
        # Si las fechas están vacías, exportar todos los pedidos
        queryset = Pedido.objects.all()

    # Agregar datos al libro de trabajo
    for row_num, pedido in enumerate(queryset, start=2):
        row = [
            pedido.pk,
            pedido.cliente.nombre if pedido.cliente else '',
            pedido.semana,
            pedido.fecha_solicitud,
            pedido.fecha_entrega,
            pedido.fecha_llegada,
            pedido.exportadora.nombre if pedido.exportadora else '',
            pedido.subexportadora.nombre if pedido.subexportadora else '',
            pedido.intermediario.nombre if pedido.intermediario else '',
            pedido.dias_cartera,
            pedido.awb,
            pedido.destino.codigo if pedido.destino else '',
            pedido.aerolinea.nombre if pedido.aerolinea else '',
            pedido.agencia_carga.nombre if pedido.agencia_carga else '',
            pedido.responsable_reserva.nombre if pedido.responsable_reserva else '',
            pedido.numero_factura,
            pedido.total_cajas_solicitadas,
            pedido.total_cajas_enviadas,
            pedido.total_peso_bruto_solicitado,
            pedido.total_peso_bruto_enviado,
            pedido.total_piezas_solicitadas,
            pedido.total_piezas_enviadas,
            pedido.peso_awb,
            pedido.eta.replace(tzinfo=None) if pedido.eta else '',
            pedido.etd.replace(tzinfo=None) if pedido.etd else '',
            pedido.variedades,
            pedido.descuento,
            pedido.nota_credito_no,
            pedido.motivo_nota_credito,
            pedido.valor_total_nota_credito_usd,
            pedido.valor_pagado_cliente_usd,
            pedido.estado_factura,
            pedido.utilidad_bancaria_usd,
            pedido.fecha_pago,
            pedido.trm_monetizacion,
            pedido.fecha_monetizacion,
            pedido.tasa_representativa_usd_diaria,
            pedido.trm_cotizacion,
            pedido.diferencia_por_abono,
            pedido.dias_de_vencimiento,
            pedido.valor_total_factura_usd,
            pedido.valor_total_utilidad_usd,
            pedido.valor_utilidad_pesos,
            pedido.documento_cobro_utilidad,
            pedido.fecha_pago_utilidad,
            pedido.estado_utilidad,
            pedido.estado_cancelacion,
            pedido.estado_documentos,
            pedido.estatus_reserva,
            pedido.termo,
            pedido.eta_real.replace(tzinfo=None) if pedido.eta_real else '',
            pedido.diferencia_peso_factura_awb,
            pedido.estado_pedido,
            pedido.observaciones_tracking,
            pedido.observaciones
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pedido_general.xlsx"'

    return response


# ------------------ Exportacion de Pedidos Excel Etnico --------------------------------------------------------

# Vista Para Exportar Pedidos Etnico:

class ExportarPedidosEtnicoView(TemplateView):
    template_name = 'export_pedidos_etnico.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Agrega contexto adicional aquí si es necesario
        return context


@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Etnico'), login_url='home'))
def exportar_pedidos_etnico(request):
    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Pedidos Totales Etnico'
    font = Font(bold=True)
    fill = PatternFill(start_color="fffaac", end_color="fffaac", fill_type="solid")

    # Encabezados
    columns = ['No', 'Cliente', 'Semana',
               'Fecha Solicitud', 'Fecha Entrega', 'Fecha Llegada', 'Exportador', 'Subexportadora', 'Intermediario',
               'Dias Cartera',
               'AWB', 'Destino', 'Aerolinea', 'Agencia De Carga', 'Responsable Reserva', 'Numero Factura',
               'Total Cajas Solicitadas', 'Total Cajas Enviadas', 'Peso Bruto Solicitado', 'Peso Bruto Enviado',
               'Pallets Solicitados', 'Pallets Enviados', 'Peso AWB', 'ETA', 'ETD', 'Variedades', 'Descuento Comercial',
               'No NC', 'Motivo NC', 'Valor Total NC', 'Valor Pagado Cliente', 'Estado Factura',
               'Utilidad Bancaria USD', 'Fecha Pago Cliente', 'TRM Monetización', 'Fecha Monetización', 'Trm Banrep',
               'Trm Cotización', 'Diferencia Pago', 'Dias Vencimiento', 'Valor Total Factura USD', 'Valor Utilidad USD',
               'Valor Utilidad Pesos', 'Documento Cobro Utilidad', 'Fecha Pago Utilidad', 'Estado Utilidad',
               'Estado Cancelacion', 'Estado Documentos', 'Estado Reserva', 'Estado Pedido', 'Observaciones Tracking',
               'Observaciones Generales']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill

    fecha_inicial_str = request.POST.get('fecha_inicial')
    fecha_final_str = request.POST.get('fecha_final')

    # Verificar si las fechas están vacías o nulas
    if fecha_inicial_str and fecha_final_str:
        # Convertir las cadenas de fecha en objetos datetime
        fecha_inicial = datetime.strptime(fecha_inicial_str, '%Y-%m-%d')
        fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d')

        # Filtrar los pedidos por fecha_entrega dentro del rango
        queryset = Pedido.objects.filter(exportadora__nombre='Etnico', fecha_entrega__gte=fecha_inicial,
                                         fecha_entrega__lte=fecha_final)
    else:
        # Si las fechas están vacías, exportar todos los pedidos de la exportadora 'Etnico'
        queryset = Pedido.objects.filter(exportadora__nombre='Etnico')

    # Agregar datos al libro de trabajo
    for row_num, pedido in enumerate(queryset, start=2):
        row = [
            pedido.pk,
            pedido.cliente.nombre if pedido.cliente else '',
            pedido.semana,
            pedido.fecha_solicitud,
            pedido.fecha_entrega,
            pedido.fecha_llegada,
            pedido.exportadora.nombre if pedido.exportadora else '',
            pedido.subexportadora.nombre if pedido.subexportadora else '',
            pedido.intermediario.nombre if pedido.intermediario else '',
            pedido.dias_cartera,
            pedido.awb,
            pedido.destino.codigo if pedido.destino else '',
            pedido.aerolinea.nombre if pedido.aerolinea else '',
            pedido.agencia_carga.nombre if pedido.agencia_carga else '',
            pedido.responsable_reserva.nombre if pedido.responsable_reserva else '',
            pedido.numero_factura,
            pedido.total_cajas_solicitadas,
            pedido.total_cajas_enviadas,
            pedido.total_peso_bruto_solicitado,
            pedido.total_peso_bruto_enviado,
            pedido.total_piezas_solicitadas,
            pedido.total_piezas_enviadas,
            pedido.peso_awb,
            pedido.eta.replace(tzinfo=None) if pedido.eta else '',
            pedido.etd.replace(tzinfo=None) if pedido.etd else '',
            pedido.variedades,
            pedido.descuento,
            pedido.nota_credito_no,
            pedido.motivo_nota_credito,
            pedido.valor_total_nota_credito_usd,
            pedido.valor_pagado_cliente_usd,
            pedido.estado_factura,
            pedido.utilidad_bancaria_usd,
            pedido.fecha_pago,
            pedido.trm_monetizacion,
            pedido.fecha_monetizacion,
            pedido.tasa_representativa_usd_diaria,
            pedido.trm_cotizacion,
            pedido.diferencia_por_abono,
            pedido.dias_de_vencimiento,
            pedido.valor_total_factura_usd,
            pedido.valor_total_utilidad_usd,
            pedido.valor_utilidad_pesos,
            pedido.documento_cobro_utilidad,
            pedido.fecha_pago_utilidad,
            pedido.estado_utilidad,
            pedido.estado_cancelacion,
            pedido.estado_documentos,
            pedido.estatus_reserva,
            pedido.estado_pedido,
            pedido.observaciones_tracking,
            pedido.observaciones
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pedidos_etnico.xlsx"'

    return response


# ------------------ Exportacion de Pedidos Excel Fieldex --------------------------------------------------------

class ExportarPedidosFieldexView(TemplateView):
    template_name = 'export_pedidos_fieldex.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Agrega contexto adicional aquí si es necesario
        return context


@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Fieldex'), login_url='home'))
def exportar_pedidos_fieldex(request):
    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Pedidos Totales Fieldex'
    font = Font(bold=True)
    fill = PatternFill(start_color="fffaac", end_color="fffaac", fill_type="solid")

    # Encabezados
    columns = ['No', 'Cliente', 'Semana',
               'Fecha Solicitud', 'Fecha Entrega', 'Fecha Llegada', 'Exportador', 'Subexportadora', 'Intermediario',
               'Dias Cartera',
               'AWB', 'Destino', 'Aerolinea', 'Agencia De Carga', 'Responsable Reserva', 'Numero Factura',
               'Total Cajas Solicitadas', 'Total Cajas Enviadas', 'Peso Bruto Solicitado', 'Peso Bruto Enviado',
               'Pallets Solicitados', 'Pallets Enviados', 'Peso AWB', 'ETA', 'ETD', 'Variedades', 'Descuento Comercial',
               'No NC', 'Motivo NC', 'Valor Total NC', 'Valor Pagado Cliente', 'Estado Factura',
               'Utilidad Bancaria USD', 'Fecha Pago Cliente', 'TRM Monetización', 'Fecha Monetización', 'Trm Banrep',
               'Trm Cotización', 'Diferencia Pago', 'Dias Vencimiento', 'Valor Total Factura USD', 'Valor Utilidad USD',
               'Valor Utilidad Pesos', 'Documento Cobro Utilidad', 'Fecha Pago Utilidad', 'Estado Utilidad',
               'Estado Cancelacion', 'Estado Documentos', 'Estado Reserva', 'Estado Pedido', 'Observaciones Tracking',
               'Observaciones Generales']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill

    fecha_inicial_str = request.POST.get('fecha_inicial')
    fecha_final_str = request.POST.get('fecha_final')

    # Verificar si las fechas están vacías o nulas
    if fecha_inicial_str and fecha_final_str:
        # Convertir las cadenas de fecha en objetos datetime
        fecha_inicial = datetime.strptime(fecha_inicial_str, '%Y-%m-%d')
        fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d')

        # Filtrar los pedidos por fecha_entrega dentro del rango
        queryset = Pedido.objects.filter(exportadora__nombre='Fieldex', fecha_entrega__gte=fecha_inicial,
                                         fecha_entrega__lte=fecha_final)
    else:
        # Si las fechas están vacías, exportar todos los pedidos de la exportadora 'Etnico'
        queryset = Pedido.objects.filter(exportadora__nombre='Fieldex')

    # Agregar datos al libro de trabajo
    for row_num, pedido in enumerate(queryset, start=2):
        row = [
            pedido.pk,
            pedido.cliente.nombre if pedido.cliente else '',
            pedido.semana,
            pedido.fecha_solicitud,
            pedido.fecha_entrega,
            pedido.fecha_llegada,
            pedido.exportadora.nombre if pedido.exportadora else '',
            pedido.subexportadora.nombre if pedido.subexportadora else '',
            pedido.intermediario.nombre if pedido.intermediario else '',
            pedido.dias_cartera,
            pedido.awb,
            pedido.destino.codigo if pedido.destino else '',
            pedido.aerolinea.nombre if pedido.aerolinea else '',
            pedido.agencia_carga.nombre if pedido.agencia_carga else '',
            pedido.responsable_reserva.nombre if pedido.responsable_reserva else '',
            pedido.numero_factura,
            pedido.total_cajas_solicitadas,
            pedido.total_cajas_enviadas,
            pedido.total_peso_bruto_solicitado,
            pedido.total_peso_bruto_enviado,
            pedido.total_piezas_solicitadas,
            pedido.total_piezas_enviadas,
            pedido.peso_awb,
            pedido.eta.replace(tzinfo=None) if pedido.eta else '',
            pedido.etd.replace(tzinfo=None) if pedido.etd else '',
            pedido.variedades,
            pedido.descuento,
            pedido.nota_credito_no,
            pedido.motivo_nota_credito,
            pedido.valor_total_nota_credito_usd,
            pedido.valor_pagado_cliente_usd,
            pedido.estado_factura,
            pedido.utilidad_bancaria_usd,
            pedido.fecha_pago,
            pedido.trm_monetizacion,
            pedido.fecha_monetizacion,
            pedido.tasa_representativa_usd_diaria,
            pedido.trm_cotizacion,
            pedido.diferencia_por_abono,
            pedido.dias_de_vencimiento,
            pedido.valor_total_factura_usd,
            pedido.valor_total_utilidad_usd,
            pedido.valor_utilidad_pesos,
            pedido.documento_cobro_utilidad,
            pedido.fecha_pago_utilidad,
            pedido.estado_utilidad,
            pedido.estado_cancelacion,
            pedido.estado_documentos,
            pedido.estatus_reserva,
            pedido.estado_pedido,
            pedido.observaciones_tracking,
            pedido.observaciones
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pedidos_fieldex.xlsx"'

    return response


# ------------------ Exportacion de Pedidos Excel Juan Matas ---------------------------------------------------------
class ExportarPedidosJuanView(TemplateView):
    template_name = 'export_pedidos_juan.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Agrega contexto adicional aquí si es necesario
        return context


@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Juan_Matas'), login_url='home'))
def exportar_pedidos_juan(request):
    # Crear un libro de trabajo de Excel
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Pedidos Totales Juan_Matas'
    font = Font(bold=True)
    fill = PatternFill(start_color="fffaac", end_color="fffaac", fill_type="solid")

    # Encabezados
    columns = ['No', 'Cliente', 'Semana',
               'Fecha Solicitud', 'Fecha Entrega', 'Fecha Llegada', 'Exportador', 'Subexportadora', 'Intermediario',
               'Dias Cartera',
               'AWB', 'Destino', 'Aerolinea', 'Agencia De Carga', 'Responsable Reserva', 'Numero Factura',
               'Total Cajas Solicitadas', 'Total Cajas Enviadas', 'Peso Bruto Solicitado', 'Peso Bruto Enviado',
               'Pallets Solicitados', 'Pallets Enviados', 'Peso AWB', 'ETA', 'ETD', 'Variedades', 'Descuento Comercial',
               'No NC', 'Motivo NC', 'Valor Total NC', 'Valor Pagado Cliente', 'Estado Factura',
               'Utilidad Bancaria USD', 'Fecha Pago Cliente', 'TRM Monetización', 'Fecha Monetización', 'Trm Banrep',
               'Trm Cotización', 'Diferencia Pago', 'Dias Vencimiento', 'Valor Total Factura USD', 'Valor Utilidad USD',
               'Valor Utilidad Pesos', 'Documento Cobro Utilidad', 'Fecha Pago Utilidad', 'Estado Utilidad',
               'Estado Cancelacion', 'Estado Documentos', 'Estado Reserva', 'Estado Pedido', 'Observaciones Tracking',
               'Observaciones Generales']

    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill

    fecha_inicial_str = request.POST.get('fecha_inicial')
    fecha_final_str = request.POST.get('fecha_final')

    # Verificar si las fechas están vacías o nulas
    if fecha_inicial_str and fecha_final_str:
        # Convertir las cadenas de fecha en objetos datetime
        fecha_inicial = datetime.strptime(fecha_inicial_str, '%Y-%m-%d')
        fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d')

        # Filtrar los pedidos por fecha_entrega dentro del rango
        queryset = Pedido.objects.filter(exportadora__nombre='Juan_Matas', fecha_entrega__gte=fecha_inicial,
                                         fecha_entrega__lte=fecha_final)
    else:
        # Si las fechas están vacías, exportar todos los pedidos de la exportadora 'Etnico'
        queryset = Pedido.objects.filter(exportadora__nombre='Juan_Matas')

    # Agregar datos al libro de trabajo
    for row_num, pedido in enumerate(queryset, start=2):
        row = [
            pedido.pk,
            pedido.cliente.nombre if pedido.cliente else '',
            pedido.semana,
            pedido.fecha_solicitud,
            pedido.fecha_entrega,
            pedido.fecha_llegada,
            pedido.exportadora.nombre if pedido.exportadora else '',
            pedido.subexportadora.nombre if pedido.subexportadora else '',
            pedido.intermediario.nombre if pedido.intermediario else '',
            pedido.dias_cartera,
            pedido.awb,
            pedido.destino.codigo if pedido.destino else '',
            pedido.aerolinea.nombre if pedido.aerolinea else '',
            pedido.agencia_carga.nombre if pedido.agencia_carga else '',
            pedido.responsable_reserva.nombre if pedido.responsable_reserva else '',
            pedido.numero_factura,
            pedido.total_cajas_solicitadas,
            pedido.total_cajas_enviadas,
            pedido.total_peso_bruto_solicitado,
            pedido.total_peso_bruto_enviado,
            pedido.total_piezas_solicitadas,
            pedido.total_piezas_enviadas,
            pedido.peso_awb,
            pedido.eta.replace(tzinfo=None) if pedido.eta else '',
            pedido.etd.replace(tzinfo=None) if pedido.etd else '',
            pedido.variedades,
            pedido.descuento,
            pedido.nota_credito_no,
            pedido.motivo_nota_credito,
            pedido.valor_total_nota_credito_usd,
            pedido.valor_pagado_cliente_usd,
            pedido.estado_factura,
            pedido.utilidad_bancaria_usd,
            pedido.fecha_pago,
            pedido.trm_monetizacion,
            pedido.fecha_monetizacion,
            pedido.tasa_representativa_usd_diaria,
            pedido.trm_cotizacion,
            pedido.diferencia_por_abono,
            pedido.dias_de_vencimiento,
            pedido.valor_total_factura_usd,
            pedido.valor_total_utilidad_usd,
            pedido.valor_utilidad_pesos,
            pedido.documento_cobro_utilidad,
            pedido.fecha_pago_utilidad,
            pedido.estado_utilidad,
            pedido.estado_cancelacion,
            pedido.estado_documentos,
            pedido.estatus_reserva,
            pedido.estado_pedido,
            pedido.observaciones_tracking,
            pedido.observaciones
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(output)
    output.seek(0)

    # Crear una respuesta HTTP con el archivo de Excel
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pedidos_juan.xlsx"'

    return response


# -------------------------- Funciones De Exportacion Cartera General--------------------------------------------------
# ------------------ Exportacion De Cartera Por cliente --------------------------------------------------------------
class ExportarCarteraClienteView(TemplateView):
    template_name = 'export_cartera_cliente.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = ExportSearchForm()
        return context

    def post(self, request, *args, **kwargs):
        form = ExportSearchForm(request.POST)
        if form.is_valid():
            cliente = form.cleaned_data.get('cliente', None)
            intermediario = form.cleaned_data.get('intermediario', None)
            fecha_inicial_str = request.POST.get('fecha_inicial', None)
            fecha_final_str = request.POST.get('fecha_final', None)

            fecha_inicial = datetime.strptime(fecha_inicial_str, "%Y-%m-%d") if fecha_inicial_str else None
            fecha_final = datetime.strptime(fecha_final_str, "%Y-%m-%d") if fecha_final_str else None

            # Determinar el grupo del usuario
            grupo = None
            if es_miembro_del_grupo('Heavens')(request.user):
                grupo = 'Heavens'
            elif es_miembro_del_grupo('Etnico')(request.user):
                grupo = 'Etnico'
            elif es_miembro_del_grupo('Fieldex')(request.user):
                grupo = 'Fieldex'
            elif es_miembro_del_grupo('Juan_Matas')(request.user):
                grupo = 'Juan_Matas'

            # Obtener los datos y los totales con el filtro de fecha, cliente, intermediario y grupo
            pedidos, totales = obtener_datos_con_totales_cliente(fecha_inicial, fecha_final, cliente, intermediario,
                                                                 grupo)

            # Crear el archivo Excel
            ruta_archivo = 'estado_cuenta_clientes.xlsx'
            crear_archivo_excel_cliente(pedidos, totales, ruta_archivo)

            # Leer el archivo y preparar la respuesta
            with open(ruta_archivo, 'rb') as archivo_excel:
                response = HttpResponse(archivo_excel.read(),
                                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="estado_cuenta_clientes.xlsx"'

            return response

        context = self.get_context_data(**kwargs)
        context['form'] = form
        return self.render_to_response(context)


# ------------------  Exportacion De Cartera Por cliente  Para Enviar --------------------------------------------

class ExportarCarteraClienteEnviarView(TemplateView):
    template_name = 'export_cartera_cliente_enviar.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = ExportSearchForm()
        return context

    def post(self, request, *args, **kwargs):
        form = ExportSearchForm(request.POST)
        if form.is_valid():
            cliente = form.cleaned_data.get('cliente', None)
            intermediario = form.cleaned_data.get('intermediario', None)
            fecha_inicial_str = request.POST.get('fecha_inicial', None)
            fecha_final_str = request.POST.get('fecha_final', None)

            fecha_inicial = datetime.strptime(fecha_inicial_str, "%Y-%m-%d") if fecha_inicial_str else None
            fecha_final = datetime.strptime(fecha_final_str, "%Y-%m-%d") if fecha_final_str else None

            # Determinar el grupo del usuario
            grupo = None
            if es_miembro_del_grupo('Heavens')(request.user):
                grupo = 'Heavens'
            elif es_miembro_del_grupo('Etnico')(request.user):
                grupo = 'Etnico'
            elif es_miembro_del_grupo('Fieldex')(request.user):
                grupo = 'Fieldex'
            elif es_miembro_del_grupo('Juan_Matas')(request.user):
                grupo = 'Juan_Matas'

            # Obtener los datos y los totales con el filtro de fecha, cliente, intermediario y grupo
            pedidos, totales = obtener_datos_con_totales_enviar_cliente(fecha_inicial, fecha_final, cliente, intermediario,
                                                                 grupo)

            # Crear el archivo Excel
            ruta_archivo = 'estado_cuenta_clientes.xlsx'
            crear_archivo_excel_enviar_cliente(pedidos, totales, ruta_archivo)

            # Leer el archivo y preparar la respuesta
            with open(ruta_archivo, 'rb') as archivo_excel:
                response = HttpResponse(archivo_excel.read(),
                                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="estado_cuenta_clientes.xlsx"'

            return response

        context = self.get_context_data(**kwargs)
        context['form'] = form
        return self.render_to_response(context)


# -------------------------------- Tabla De Pedidos General  ----------------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoListView(SingleTableView):
    model = Pedido
    table_class = PedidoTable
    table_pagination = {"per_page": 14}
    template_name = 'pedido_list_general.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().prefetch_related('autorizacioncancelacion_set')
        form = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')
            cliente = form.cleaned_data.get('cliente')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()
                elif metodo_busqueda == 'intermediario':
                    queryset = queryset.filter(intermediario__nombre__icontains=item_busqueda)

            if cliente:
                queryset = queryset.filter(cliente=cliente)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())
        context['es_autorizador'] = self.request.user.groups.filter(name='Autorizadores').exists()
        return context

    def get_clientes_con_pedidos(self):
        return Cliente.objects.filter(pedido__isnull=False).distinct()


# -------------------------------- Tabla De Pedidos Seguimientos ----------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class SeguimientosPedidosListView(SingleTableView):
    model = Pedido
    table_class = SeguimienosTable
    table_pagination = {"per_page": 14}
    template_name = 'seguimiento_pedido_list_general.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().prefetch_related('autorizacioncancelacion_set')
        form = self.form_class(self.request.GET)

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'cliente':
                    queryset = queryset.filter(cliente__nombre__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()  # No results if ID is not an integer

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET)
        context['es_autorizador'] = self.request.user.groups.filter(name='Autorizadores').exists()
        return context


# -------------------------------- Tabla De Pedidos Seguimientos Resumen ----------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class ResumenSeguimientosPedidosListView(SingleTableView):
    model = Pedido
    table_class = SeguimienosResumenTable
    table_pagination = {"per_page": 50}
    template_name = 'resumen_seguimiento_pedido_list_general.html'
    form_class = FiltroSemanaExportadoraForm

    def get_queryset(self):
        queryset = super().get_queryset().prefetch_related('autorizacioncancelacion_set')
        form = self.form_class(self.request.GET)

        if form.is_valid():
            semana = form.cleaned_data.get('semana')
            exportadora = form.cleaned_data.get('exportadora')

            if semana:
                queryset = queryset.filter(semana=semana)
            if exportadora:
                queryset = queryset.filter(exportadora=exportadora)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET)
        context['es_autorizador'] = self.request.user.groups.filter(name='Autorizadores').exists()
        return context


# -------------------------------- Tabla De Pedidos Etnico  ----------------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Etnico'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoEtnicoListView(SingleTableView):
    model = Pedido
    table_class = PedidoExportadorTable
    table_pagination = {"per_page": 14}
    template_name = 'pedido_list_etnico.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Etnico')
        form = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')
            cliente = form.cleaned_data.get('cliente')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()
                elif metodo_busqueda == 'intermediario':
                    queryset = queryset.filter(intermediario__nombre__icontains=item_busqueda)

            if cliente:
                queryset = queryset.filter(cliente=cliente)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())
        return context

    def get_clientes_con_pedidos(self):
        return Cliente.objects.filter(pedido__isnull=False).distinct()


# -------------------------------- Tabla De Pedidos Fieldex  ----------------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Fieldex'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoFieldexListView(SingleTableView):
    model = Pedido
    table_class = PedidoExportadorTable
    table_pagination = {"per_page": 14}
    template_name = 'pedido_list_fieldex.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Fieldex')
        form = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')
            cliente = form.cleaned_data.get('cliente')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()
                elif metodo_busqueda == 'intermediario':
                    queryset = queryset.filter(intermediario__nombre__icontains=item_busqueda)

            if cliente:
                queryset = queryset.filter(cliente=cliente)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())
        return context

    def get_clientes_con_pedidos(self):
        return Cliente.objects.filter(pedido__isnull=False).distinct()


# -------------------------------- Tabla De Pedidos Juan Matas  ----------------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Juan_Matas'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoJuanListView(SingleTableView):
    model = Pedido
    table_class = PedidoExportadorTable
    table_pagination = {"per_page": 14}
    template_name = 'pedido_list_Juan.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Juan_Matas')
        form = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')
            cliente = form.cleaned_data.get('cliente')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()
                elif metodo_busqueda == 'intermediario':
                    queryset = queryset.filter(intermediario__nombre__icontains=item_busqueda)

            if cliente:
                queryset = queryset.filter(cliente=cliente)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())
        return context

    def get_clientes_con_pedidos(self):
        return Cliente.objects.filter(pedido__isnull=False).distinct()


# -------------------------------  Formulario - Crear Pedido General - Modal (General) ----------------------------
@method_decorator(login_required, name='dispatch')
class PedidoCreateView(CreateView):
    model = Pedido
    form_class = PedidoForm
    template_name = 'pedido_crear.html'
    success_url = '/pedido_list_general/'

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        return form

    @transaction.atomic
    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request,
                         f"El pedido para el cliente {form.cleaned_data['cliente']} se ha creado exitosamente.")
        return JsonResponse({'success': True})

    def form_invalid(self, form):
        return JsonResponse({'success': False, 'html': render_to_string(self.template_name, {'form': form})})


# -------------------------------  Formulario - Editar Pedido General - Modal (General) ----------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoUpdateView(UpdateView):
    model = Pedido
    form_class = EditarPedidoForm
    template_name = 'pedido_editar.html'
    success_url = '/pedido_list_general/'

    def get_object(self, queryset=None):
        pedido_id = int(self.request.POST.get('pedido_id').replace(".", "")) if self.request.POST.get(
            'pedido_id') else int(self.request.GET.get('pedido_id').replace(".", ""))
        pedido = get_object_or_404(Pedido, id=pedido_id)
        return pedido

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido'] = self.object
        return context

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()

        formatted_fecha_entrega = self.object.fecha_entrega.strftime('%Y-%m-%d') if self.object.fecha_entrega else ''
        formatted_fecha_pago_utilidad = self.object.fecha_pago_utilidad.strftime(
            '%Y-%m-%d') if self.object.fecha_pago_utilidad else ''

        form = self.form_class(
            instance=self.object,
            initial={'fecha_entrega': formatted_fecha_entrega, 'fecha_pago_utilidad': formatted_fecha_pago_utilidad}
        )

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form, 'pedido': self.object}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return self.render_to_response(self.get_context_data(form=form))

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['instance'] = self.object
        return kwargs

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, f"El pedido ha sido editado exitosamente.")
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False,
                                 'html': render_to_string(self.template_name, {'form': form, 'pedido': self.object},
                                                          request=self.request)})
        else:
            return super().form_invalid(form)


# -------------------------------  Formulario - Editar Pedido Segunda parte - Modal (General) ----------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoUpdateViewDos(UpdateView):
    model = Pedido
    form_class = EditarPedidoFormDos
    template_name = 'pedido_editar2.html'
    success_url = '/pedido_list_general/'

    def get_object(self, queryset=None):
        pedido_id = int(self.request.POST.get('pedido_id').replace(".", "")) if self.request.POST.get(
            'pedido_id') else int(self.request.GET.get('pedido_id').replace(".", ""))
        pedido = get_object_or_404(Pedido, id=pedido_id)
        return pedido

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido'] = self.object
        return context

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()

        formatted_fecha_entrega = self.object.fecha_entrega.strftime('%Y-%m-%d') if self.object.fecha_entrega else ''
        formatted_fecha_pago_utilidad = self.object.fecha_pago_utilidad.strftime(
            '%Y-%m-%d') if self.object.fecha_pago_utilidad else ''

        form = self.form_class(
            instance=self.object,
            initial={'fecha_entrega': formatted_fecha_entrega, 'fecha_pago_utilidad': formatted_fecha_pago_utilidad}
        )

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form, 'pedido': self.object}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return self.render_to_response(self.get_context_data(form=form))

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['instance'] = self.object
        return kwargs

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, f"El pedido ha sido editado exitosamente.")
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False,
                                 'html': render_to_string(self.template_name, {'form': form, 'pedido': self.object},
                                                          request=self.request)})
        else:
            return super().form_invalid(form)


# -------------------------------  Formulario - Editar Pedido Cartera - Modal (General) ----------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoUpdateViewCartera(UpdateView):
    model = Pedido
    form_class = EditarPedidoFormCartera
    template_name = 'pedido_editar_cartera.html'
    success_url = '/cartera_list_heavens'

    def get_object(self, queryset=None):
        pedido_id = int(self.request.POST.get('pedido_id').replace(".", "")) if self.request.POST.get(
            'pedido_id') else int(self.request.GET.get('pedido_id').replace(".", ""))
        pedido = get_object_or_404(Pedido, id=pedido_id)
        return pedido

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido'] = self.object
        return context

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()

        formatted_fecha_entrega = self.object.fecha_entrega.strftime('%Y-%m-%d') if self.object.fecha_entrega else ''
        formatted_fecha_pago_utilidad = self.object.fecha_pago_utilidad.strftime(
            '%Y-%m-%d') if self.object.fecha_pago_utilidad else ''

        form = self.form_class(
            instance=self.object,
            initial={'fecha_entrega': formatted_fecha_entrega, 'fecha_pago_utilidad': formatted_fecha_pago_utilidad}
        )

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form, 'pedido': self.object}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return self.render_to_response(self.get_context_data(form=form))

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['instance'] = self.object
        return kwargs

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, f"El pedido ha sido editado exitosamente.")
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False,
                                 'html': render_to_string(self.template_name, {'form': form, 'pedido': self.object},
                                                          request=self.request)})
        else:
            return super().form_invalid(form)


# --------------------------------------- Editar pedido Utilidades -------------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoUpdateViewUtilidades(UpdateView):
    model = Pedido
    form_class = EditarPedidoFormUtilidades
    template_name = 'pedido_editar_utilidades.html'
    success_url = '/utilidad_list_heavens/'

    def get_object(self, queryset=None):
        pedido_id = int(self.request.POST.get('pedido_id').replace(".", "")) if self.request.POST.get(
            'pedido_id') else int(self.request.GET.get('pedido_id').replace(".", ""))
        pedido = get_object_or_404(Pedido, id=pedido_id)
        return pedido

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido'] = self.object
        return context

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()

        formatted_fecha_entrega = self.object.fecha_entrega.strftime('%Y-%m-%d') if self.object.fecha_entrega else ''
        formatted_fecha_pago_utilidad = self.object.fecha_pago_utilidad.strftime(
            '%Y-%m-%d') if self.object.fecha_pago_utilidad else ''

        form = self.form_class(
            instance=self.object,
            initial={'fecha_entrega': formatted_fecha_entrega, 'fecha_pago_utilidad': formatted_fecha_pago_utilidad}
        )

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form, 'pedido': self.object}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return self.render_to_response(self.get_context_data(form=form))

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['instance'] = self.object
        return kwargs

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, f"El pedido ha sido editado exitosamente.")
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False,
                                 'html': render_to_string(self.template_name, {'form': form, 'pedido': self.object},
                                                          request=self.request)})
        else:
            return super().form_invalid(form)


# -------------------------------  //// Formulario - Editar Pedido Por Exportador //// ----------------------------
@method_decorator(login_required, name='dispatch')
class PedidoExportadorUpdateView(UpdateView):
    model = Pedido
    form_class = EditarPedidoExportadorForm
    template_name = 'pedido_editar_exportador.html'
    success_url = '/update_items/'

    def get_object(self, queryset=None):
        pedido_id = int(self.request.POST.get('pedido_id').replace(".", "")) if self.request.POST.get(
            'pedido_id') else int(self.request.GET.get('pedido_id').replace(".", ""))
        pedido = get_object_or_404(Pedido, id=pedido_id)
        return pedido

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido'] = self.object
        return context

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()

        formatted_fecha_monetizacion = self.object.fecha_monetizacion.strftime(
            '%Y-%m-%d') if self.object.fecha_monetizacion else ''
        formatted_fecha_pago = self.object.fecha_pago.strftime('%Y-%m-%d') if self.object.fecha_pago else ''

        form = self.form_class(
            instance=self.object,
            initial={'fecha_monetizacion': formatted_fecha_monetizacion, 'fecha_pago': formatted_fecha_pago}
        )

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form, 'pedido': self.object}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return self.render_to_response(self.get_context_data(form=form))

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, f"El pedido ha sido editado exitosamente.")
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False,
                                 'html': render_to_string(self.template_name, {'form': form, 'pedido': self.object},
                                                          request=self.request)})
        else:
            return super().form_invalid(form)


# -------------------------------  Formulario - Editar Pedido Seguimiento - Modal (Tracking)----------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoUpdateSebguimientoView(UpdateView):
    model = Pedido
    form_class = EditarPedidoSeguimientoForm
    template_name = 'pedido_editar_seguimiento.html'
    success_url = '/seguimiento_pedido_list_general/'

    def get_object(self, queryset=None):
        pedido_id = int(self.request.POST.get('pedido_id').replace(".", "")) if self.request.POST.get(
            'pedido_id') else int(self.request.GET.get('pedido_id').replace(".", ""))
        pedido = get_object_or_404(Pedido, id=pedido_id)
        return pedido

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido'] = self.object
        return context

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        formatted_fecha_llegada = self.object.etd.strftime('%Y-%m-%d') if self.object.etd else ''
        formatted_etd = self.object.etd.strftime('%Y-%m-%dT%H:%M') if self.object.etd else ''
        formatted_eta = self.object.eta.strftime('%Y-%m-%dT%H:%M') if self.object.eta else ''
        formatted_eta_real = self.object.eta_real.strftime('%Y-%m-%dT%H:%M') if self.object.eta_real else ''

        form = self.form_class(
            instance=self.object,
            initial={'fecha_llegada': formatted_fecha_llegada, 'etd': formatted_etd, 'eta': formatted_eta,
                     'eta_real': formatted_eta_real})

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form, 'pedido': self.object}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return self.render_to_response(self.get_context_data(form=form))

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, f"El pedido ha sido editado exitosamente.")
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False,
                                 'html': render_to_string(self.template_name, {'form': form, 'pedido': self.object},
                                                          request=self.request)})
        else:
            return super().form_invalid(form)


# -------------------------------  Formulario - Eliminar Pedido General - Modal (General) ----------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class PedidoDeleteView(UpdateView):
    model = Pedido
    form_class = EliminarPedidoForm
    template_name = 'pedido_eliminar.html'
    success_url = '/pedido_list_general/'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.object = None

    def get_object(self, queryset=None):
        pedido_id = int(self.request.POST.get('pedido_id').replace(".", ""))
        pedido = get_object_or_404(Pedido, id=pedido_id)
        return pedido

    def get(self, request, *args, **kwargs):
        pedido_id = int(request.GET.get('pedido_id').replace(".", ""))
        self.object = get_object_or_404(Pedido, id=pedido_id)
        formatted_fecha_solicitud = self.object.fecha_solicitud.strftime('%Y-%m-%d')
        formatted_fecha_entrega = self.object.fecha_entrega.strftime('%Y-%m-%d')
        form = self.form_class(
            instance=self.object,
            initial={'fecha_solicitud': formatted_fecha_solicitud, 'fecha_entrega': formatted_fecha_entrega}
        )
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        pedido = form.save(commit=False)
        pedido.delete()
        messages.warning(self.request,
                         f"El pedido para el cliente {form.cleaned_data['cliente']} se ha eliminado exitosamente.")
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse(
                {'success': False, 'html': render_to_string(self.template_name, {'form': form}, request=self.request)})
        else:
            return super().form_invalid(form)


# ----------------------------------- Mostrar Detalles De Pedido General ------------------------------------------
@method_decorator(login_required, name='dispatch')
class DetallePedidoListView(SingleTableView):
    model = DetallePedido
    table_class = DetallePedidoTable
    template_name = 'pedido_detalle_list.html'

    def dispatch(self, request, *args, **kwargs):
        pedido_id = self.kwargs.get('pedido_id')
        pedido = get_object_or_404(Pedido, pk=pedido_id)

        # Comprueba si el usuario pertenece al grupo requerido
        if not request.user.groups.filter(name=pedido.exportadora.nombre).exists():
            return HttpResponseForbidden("No tienes permiso para ver estos detalles del pedido")

        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        pedido_id = self.kwargs.get('pedido_id')
        queryset = DetallePedido.objects.filter(pedido__id=pedido_id)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pedido_id = self.kwargs.get('pedido_id')
        context['pedido'] = get_object_or_404(Pedido, pk=pedido_id)
        return context


# --------------------------- Formulario Crear  Detalle De Pedido ----------------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class DetallePedidoCreateView(CreateView):
    model = DetallePedido
    form_class = DetallePedidoForm
    template_name = 'detalle_pedido_crear.html'
    success_url = '/detalle_pedido_crear/'

    def get_initial(self):
        initial = super().get_initial()
        pedido_id = self.kwargs.get('pedido_id') or self.request.GET.get('pedido_id')
        if pedido_id:
            initial['pedido'] = pedido_id
        return initial

    def get_form_kwargs(self):
        kwargs = super(DetallePedidoCreateView, self).get_form_kwargs()
        kwargs['pedido_id'] = self.kwargs.get('pedido_id')
        return kwargs

    @transaction.atomic
    def form_valid(self, form):
        pedido_id = self.kwargs.get('pedido_id')
        if pedido_id:
            pedido = get_object_or_404(Pedido, pk=pedido_id)
            form.instance.pedido = pedido

        self.object = form.save()
        messages.success(self.request, f'El detalle de pedido para el pedido {pedido_id} se ha creado exitosamente.')
        return JsonResponse({'success': True})

    def form_invalid(self, form):
        errors = form.errors.as_json()
        return JsonResponse({'success': False, 'error': errors})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido_id'] = self.kwargs.get('pedido_id')
        return context


# ---------------------------- Formulario Editar Detalle De Pedido --------------------------------------------

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class DetallePedidoUpdateView(UpdateView):
    model = DetallePedido
    form_class = EditarDetallePedidoForm
    template_name = 'detalle_pedido_editar.html'
    success_url = reverse_lazy('pedido_detalle_list')

    def get_object(self, queryset=None):
        detallepedido_id = self.request.GET.get('detallepedido_id') or self.request.POST.get('detallepedido_id')
        if not detallepedido_id:
            raise ValueError("No se proporcionó 'detallepedido_id'")
        return get_object_or_404(DetallePedido, id=int(detallepedido_id))

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        pedido_id = request.GET.get('pedido_id')
        form = self.form_class(instance=self.object, pedido_id=pedido_id)
        context = self.get_context_data(form=form)

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, context, request=request)
            return JsonResponse({'form': form_html})
        else:
            return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        pedido_id = request.POST.get('pedido_id')
        form = self.form_class(request.POST, instance=self.object, pedido_id=pedido_id)
        context = self.get_context_data(form=form)

        if form.is_valid():
            return self.form_valid(form, pedido_id)
        else:
            return self.form_invalid(form, context)

    @transaction.atomic
    def form_valid(self, form, pedido_id):
        self.object = form.save()
        messages.success(self.request,
                         f"El detalle para el pedido {pedido_id} se ha editado exitosamente.")
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form, context):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, context, request=self.request)
            return JsonResponse({'success': False, 'form_html': form_html})
        else:
            return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido_id'] = self.request.GET.get('pedido_id') or self.request.POST.get('pedido_id')
        return context


# --------------------------- Formulario Editar Momento 2 Detalle De Pedido ------------------------------------------

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class DetallePedidoUpdateDosView(UpdateView):
    model = DetallePedido
    form_class = EditarDetallePedidoDosForm
    template_name = 'detalle_pedido_editar2.html'
    success_url = reverse_lazy('pedido_detalle_list')

    def get_object(self, queryset=None):
        detallepedido_id = self.request.GET.get('detallepedido_id') or self.request.POST.get('detallepedido_id')
        if not detallepedido_id:
            raise ValueError("No se proporcionó 'detallepedido_id'")
        return get_object_or_404(DetallePedido, id=int(detallepedido_id))

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        pedido_id = request.GET.get('pedido_id')
        form = self.form_class(instance=self.object, pedido_id=pedido_id)
        context = self.get_context_data(form=form)

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, context, request=request)
            return JsonResponse({'form': form_html})
        else:
            return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        pedido_id = request.POST.get('pedido_id')
        form = self.form_class(request.POST, instance=self.object, pedido_id=pedido_id)
        context = self.get_context_data(form=form)

        if form.is_valid():
            return self.form_valid(form, pedido_id)
        else:
            return self.form_invalid(form, context)

    @transaction.atomic
    def form_valid(self, form, pedido_id):
        self.object = form.save()
        messages.success(self.request,
                         f"El detalle para el pedido {pedido_id} se ha editado exitosamente.")
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form, context):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, context, request=self.request)
            return JsonResponse({'success': False, 'form_html': form_html})
        else:
            return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido_id'] = self.request.GET.get('pedido_id') or self.request.POST.get('pedido_id')
        return context


# --------------------------- Formulario Editar Momento 3 Detalle De Pedido ------------------------------------------

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class DetallePedidoUpdateTresView(UpdateView):
    model = DetallePedido
    form_class = EditarDetallePedidoTresForm
    template_name = 'detalle_pedido_editar3.html'
    success_url = reverse_lazy('pedido_detalle_list')

    def get_object(self, queryset=None):
        detallepedido_id = self.request.GET.get('detallepedido_id') or self.request.POST.get('detallepedido_id')
        if not detallepedido_id:
            raise ValueError("No se proporcionó 'detallepedido_id'")
        return get_object_or_404(DetallePedido, id=int(detallepedido_id))

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        pedido_id = request.GET.get('pedido_id')
        form = self.form_class(instance=self.object, pedido_id=pedido_id)
        context = self.get_context_data(form=form)

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, context, request=request)
            return JsonResponse({'form': form_html})
        else:
            return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        pedido_id = request.POST.get('pedido_id')
        form = self.form_class(request.POST, instance=self.object, pedido_id=pedido_id)
        context = self.get_context_data(form=form)

        if form.is_valid():
            return self.form_valid(form, pedido_id)
        else:
            return self.form_invalid(form, context)

    @transaction.atomic
    def form_valid(self, form, pedido_id):
        self.object = form.save()
        messages.success(self.request,
                         f"El detalle para el pedido {pedido_id} se ha editado exitosamente.")
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form, context):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, context, request=self.request)
            return JsonResponse({'success': False, 'form_html': form_html})
        else:
            return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pedido_id'] = self.request.GET.get('pedido_id') or self.request.POST.get('pedido_id')
        return context


# ---------------------------- Formulario Eliminar Detalle De Pedido --------------------------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class DetallePedidoDeleteiew(UpdateView):
    model = DetallePedido
    form_class = EliminarDetallePedidoForm
    template_name = 'detalle_pedido_eliminart.html'
    success_url = '/detalle_pedido_eliminar/'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.object = None

    def get_object(self, queryset=None):
        detallepedido_id = int(self.request.POST.get('detallepedido_id').replace(".", ""))
        detallepedido = get_object_or_404(DetallePedido, id=detallepedido_id)
        return detallepedido

    def get(self, request, *args, **kwargs):
        detallepedido_id = int(request.GET.get('detallepedido_id').replace(".", ""))
        self.object = get_object_or_404(DetallePedido, id=detallepedido_id)
        form = self.form_class(
            instance=self.object,
        )
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return super().get(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['instance'] = self.object
        return kwargs

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        detallepedido = form.save(commit=False)
        detallepedido.delete()
        messages.warning(self.request,
                         f"El detalle de {form.cleaned_data['pedido']} se ha eliminado exitosamente.")
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse(
                {'success': False, 'html': render_to_string(self.template_name, {'form': form}, request=self.request)})
        else:
            return super().form_invalid(form)


# ------------------------- CARTERA General /// Table mostrar cartera de pedidos Heavens ///  -------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class CarteraHeavensListView(SingleTableView):
    model = Pedido
    table_class = CarteraPedidoTable
    table_pagination = {"per_page": 14}
    template_name = 'cartera_list_heavens.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset()
        form = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')
            cliente = form.cleaned_data.get('cliente')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()
                elif metodo_busqueda == 'intermediario':
                    queryset = queryset.filter(intermediario__nombre__icontains=item_busqueda)

            if cliente:
                queryset = queryset.filter(cliente=cliente)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET, clientes=self.get_clientes_con_pedidos())
        return context

    def get_clientes_con_pedidos(self):
        return Cliente.objects.filter(pedido__isnull=False).distinct()


# ------------------------- CARTERA /// Table mostrar cartera de pedidos Etnico ///  ---------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Etnico'), login_url=reverse_lazy('home')), name='dispatch')
class CarteraEtnicoListView(SingleTableView):
    model = Pedido
    table_class = CarteraPedidoTable
    table_pagination = {"per_page": 14}
    template_name = 'cartera_list_etnico.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Etnico')
        form = self.form_class(self.request.GET)

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'cliente':
                    queryset = queryset.filter(cliente__nombre__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# ------------------------- CARTERA /// Table mostrar cartera de pedidos Fieldex ///  ---------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Fieldex'), login_url=reverse_lazy('home')), name='dispatch')
class CarteraFieldexListView(SingleTableView):
    model = Pedido
    table_class = CarteraPedidoTable
    table_pagination = {"per_page": 14}
    template_name = 'cartera_list_fieldex.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Fieldex')
        form = self.form_class(self.request.GET)

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'cliente':
                    queryset = queryset.filter(cliente__nombre__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# ------------------------- CARTERA /// Table mostrar cartera de pedidos Juan Matas ///  ---------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Juan_Matas'), login_url=reverse_lazy('home')), name='dispatch')
class CarteraJuanListView(SingleTableView):
    model = Pedido
    table_class = CarteraPedidoTable
    table_pagination = {"per_page": 14}
    template_name = 'cartera_list_juan.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Juan_Matas')
        form = self.form_class(self.request.GET)

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'cliente':
                    queryset = queryset.filter(cliente__nombre__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# ------------------------- Utilidades GENERAL /// Table mostrar Utilidades de pedidos Heavens ///  ------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home')), name='dispatch')
class UtilidadHeavensListView(SingleTableView):
    model = Pedido
    table_class = UtilidadPedidoTable
    table_pagination = {"per_page": 14}
    template_name = 'utilidad_list_heavens.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset()
        form = self.form_class(self.request.GET)

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'cliente':
                    queryset = queryset.filter(cliente__nombre__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# ------------------------- Utilidades /// Table mostrar Utilidades de pedidos Etnico ///  ---------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Etnico'), login_url=reverse_lazy('home')), name='dispatch')
class UtilidadEtnicoListView(SingleTableView):
    model = Pedido
    table_class = UtilidadPedidoTable
    table_pagination = {"per_page": 14}
    template_name = 'utilidad_list_etnico.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Etnico')
        form = self.form_class(self.request.GET)

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'cliente':
                    queryset = queryset.filter(cliente__nombre__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# ------------------------- Utilidades /// Table mostrar Utilidades de pedidos Fieldex ///  ---------------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Fieldex'), login_url=reverse_lazy('home')), name='dispatch')
class UtilidadFiedexListView(SingleTableView):
    model = Pedido
    table_class = UtilidadPedidoTable
    table_pagination = {"per_page": 14}
    template_name = 'utilidad_list_fieldex.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Fieldex')
        form = self.form_class(self.request.GET)

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'cliente':
                    queryset = queryset.filter(cliente__nombre__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# ------------------------- Utilidades /// Table mostrar Utilidades de pedidos Juan Matas // -----------------------
@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Juan_Matas'), login_url=reverse_lazy('home')), name='dispatch')
class UtilidadJuanListView(SingleTableView):
    model = Pedido
    table_class = UtilidadPedidoTable
    table_pagination = {"per_page": 14}
    template_name = 'utilidad_list_juan.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportadora__nombre='Juan_Matas')
        form = self.form_class(self.request.GET)

        if form.is_valid():
            metodo_busqueda = form.cleaned_data.get('metodo_busqueda')
            item_busqueda = form.cleaned_data.get('item_busqueda')

            if metodo_busqueda and item_busqueda:
                if metodo_busqueda == 'awb':
                    queryset = queryset.filter(awb__icontains=item_busqueda)
                elif metodo_busqueda == 'numero_factura':
                    queryset = queryset.filter(numero_factura__icontains=item_busqueda)
                elif metodo_busqueda == 'cliente':
                    queryset = queryset.filter(cliente__nombre__icontains=item_busqueda)
                elif metodo_busqueda == 'id':
                    try:
                        item_busqueda_id = int(item_busqueda)
                        queryset = queryset.filter(id=item_busqueda_id)
                    except ValueError:
                        queryset = queryset.none()

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item_busqueda'] = self.form_class(self.request.GET)
        return context


# --------------------------------- Referencias Table Etnico----------------------------------------------------------

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Etnico'), login_url=reverse_lazy('home')), name='dispatch')
class ReferenciasEtnicoListView(SingleTableView):
    model = Referencias
    table_class = ReferenciasTable
    template_name = 'referencia_list_etnico.html'
    form_class = SearchFormReferencias

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportador__nombre='Etnico')

        form = self.form_class(self.request.GET)
        if form.is_valid():
            item_busqueda = form.cleaned_data.get('item_busqueda')
            if item_busqueda:
                queryset = queryset.filter(nombre__icontains=item_busqueda)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET)
        return context


# --------------------------------- Referencias Table Fieldex----------------------------------------------------------

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Fieldex'), login_url=reverse_lazy('home')), name='dispatch')
class ReferenciasFieldexListView(SingleTableView):
    model = Referencias
    table_class = ReferenciasTable
    template_name = 'referencia_list_fieldex.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportador__nombre='Fieldex')

        form = self.form_class(self.request.GET)
        if form.is_valid():
            item_busqueda = form.cleaned_data.get('item_busqueda')
            if item_busqueda:
                queryset = queryset.filter(nombre__icontains=item_busqueda)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET)
        return context


# --------------------------------- Referencias Table Juan_Matas ------------------------------------------------------

@method_decorator(login_required, name='dispatch')
@method_decorator(user_passes_test(es_miembro_del_grupo('Juan_Matas'), login_url=reverse_lazy('home')), name='dispatch')
class ReferenciasjuanListView(SingleTableView):
    model = Referencias
    table_class = ReferenciasTable
    template_name = 'referencia_list_juan.html'
    form_class = SearchForm

    def get_queryset(self):
        queryset = super().get_queryset().filter(exportador__nombre='Juan_Matas')

        form = self.form_class(self.request.GET)
        if form.is_valid():
            item_busqueda = form.cleaned_data.get('item_busqueda')
            if item_busqueda:
                queryset = queryset.filter(nombre__icontains=item_busqueda)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = self.form_class(self.request.GET)
        return context


# ---------------------------- Formulario Editar Referencia -----------------------------------------------------------
@method_decorator(login_required, name='dispatch')
class ReferenciaUpdateView(UpdateView):
    model = Referencias
    form_class = EditarReferenciaForm
    template_name = 'referencia_editar.html'
    success_url = reverse_lazy('referencia_editar')

    def get_object(self, queryset=None):
        referencia_id = int(self.request.POST.get('referencia_id').replace(".", ""))
        referencia = get_object_or_404(Referencias, id=referencia_id)
        return referencia

    def get(self, request, *args, **kwargs):
        referencia_id = int(request.GET.get('referencia_id').replace(".", ""))
        self.object = get_object_or_404(Referencias, id=referencia_id)
        form = self.form_class(instance=self.object)
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form_html = render_to_string(self.template_name, {'form': form}, request=request)
            return JsonResponse({'form': form_html})
        else:
            return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    @transaction.atomic
    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, f'La referencia {self.object.nombre} se ha editado exitosamente.')
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse(
                {'success': False, 'html': render_to_string(self.template_name, {'form': form}, request=self.request)})
        else:
            return super().form_invalid(form)


# ----------------------- Exportar Referencias por exportador ETNICO -------------------------------------------------

@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Etnico'), login_url='home'))
def exportar_referencias_etnico(request):
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Referencias Etnico'
    font = Font(bold=True)
    fill = PatternFill(start_color="fffaac", end_color="fffaac", fill_type="solid")

    # Encabezados
    columns = ['Referencia', 'Referencia Nueva', 'Contenedor', 'Cantidad Contendedor', 'Precio', 'Exportador']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill

    queryset = Referencias.objects.filter(exportador__nombre='Etnico')

    # Agregar datos al libro de trabajo
    for row_num, referencia in enumerate(queryset, start=2):
        contenedor_nombre = referencia.contenedor.nombre if referencia.contenedor else 'Sin Contenedor'
        row = [
            referencia.nombre,
            referencia.referencia_nueva,
            contenedor_nombre,
            referencia.cant_contenedor,
            referencia.precio,
            referencia.exportador.nombre,
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="referencias_etnico.xlsx"'

    return response


# ----------------------- Exportar Referencias por exportador FIELDEX -------------------------------------------------

@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Fieldex'), login_url='home'))
def exportar_referencias_fieldex(request):
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Referencias Fieldex'
    font = Font(bold=True)
    fill = PatternFill(start_color="fffaac", end_color="fffaac", fill_type="solid")

    # Encabezados
    columns = ['Referencia', 'Referencia Nueva', 'Contenedor', 'Cantidad Contendedor', 'Precio', 'Exportador']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill

    queryset = Referencias.objects.filter(exportador__nombre='Fieldex')

    # Agregar datos al libro de trabajo
    for row_num, referencia in enumerate(queryset, start=2):
        contenedor_nombre = referencia.contenedor.nombre if referencia.contenedor else 'Sin Contenedor'
        row = [
            referencia.nombre,
            referencia.referencia_nueva,
            contenedor_nombre,
            referencia.cant_contenedor,
            referencia.precio,
            referencia.exportador.nombre,
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="referencias_fieldex.xlsx"'

    return response


# ----------------------- Exportar Referencias por exportador JUAN -------------------------------------------------

@login_required
@user_passes_test(user_passes_test(es_miembro_del_grupo('Juan_Matas'), login_url='home'))
def exportar_referencias_juan(request):
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Referencias Juan_Matas'
    font = Font(bold=True)
    fill = PatternFill(start_color="fffaac", end_color="fffaac", fill_type="solid")

    # Encabezados
    columns = ['Referencia', 'Referencia Nueva', 'Contenedor', 'Cantidad Contendedor', 'Precio', 'Exportador']
    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = font
        cell.fill = fill

    queryset = Referencias.objects.filter(exportador__nombre='Juan_Matas')

    # Agregar datos al libro de trabajo
    for row_num, referencia in enumerate(queryset, start=2):
        contenedor_nombre = referencia.contenedor.nombre if referencia.contenedor else 'Sin Contenedor'
        row = [
            referencia.nombre,
            referencia.referencia_nueva,
            contenedor_nombre,
            referencia.cant_contenedor,
            referencia.precio,
            referencia.exportador.nombre,
        ]
        for col_num, cell_value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="referencias_juan.xlsx"'

    return response


# --------------------------------- Funciones Heavens General ---------------------------------------------------------
# ----------------------------- Actualizar los días de Vencimiento ----------------------------------------------------

def actualizar_dias_de_vencimiento_todos(request):
    pedidos = Pedido.objects.all()
    for pedido in pedidos:
        pedido.save()  # Esto llamará a tu método save personalizado
    messages.success(request, "Todos los pedidos se han actualizado correctamente.")
    return redirect('pedido_list_general')


# ------------------------ Vista para actualizar la TRM Banco de la republica -----------------------------------------
def actualizar_tasas(request):
    pedidos = Pedido.objects.order_by('-id')[:50]
    for pedido in pedidos:
        pedido.actualizar_tasa_representativa()
    messages.success(request, 'Se Actualizaron Las Tasas Con Banco De La Republica Correctamente')
    return redirect('pedido_list_general')


# --------------------------------- Funciones Etnico  ---------------------------------------------------------
def actualizar_dias_de_vencimiento_etnico(request):
    pedidos = Pedido.objects.all()
    for pedido in pedidos:
        pedido.actualizar_dias_de_vencimiento()
    messages.success(request, 'Días de vencimiento actualizados para todos los pedidos.')
    return redirect('pedido_list_etnico')


def actualizar_tasas_etnico(request):
    pedidos = Pedido.objects.order_by('-id')[:50]
    for pedido in pedidos:
        pedido.actualizar_tasa_representativa()
    messages.success(request, 'Se Actualizo La TRM Con Banco De La Republica Correctamente')
    return redirect('pedido_list_etnico')


# --------------------------------- Funciones Fieldex  ---------------------------------------------------------
def actualizar_dias_de_vencimiento_fieldex(request):
    pedidos = Pedido.objects.all()
    for pedido in pedidos:
        pedido.actualizar_dias_de_vencimiento()
    messages.success(request, 'Días de vencimiento actualizados para todos los pedidos.')
    return redirect('pedido_list_fieldex')


def actualizar_tasas_fieldex(request):
    pedidos = Pedido.objects.order_by('-id')[:50]
    for pedido in pedidos:
        pedido.actualizar_tasa_representativa()
    messages.success(request, 'Se Actualizo La TRM Con Banco De La Republica Correctamente')
    return redirect('pedido_list_fieldex')


# --------------------------------- Funciones Juan_matas  ---------------------------------------------------------
def actualizar_dias_de_vencimiento_juan(request):
    pedidos = Pedido.objects.all()
    for pedido in pedidos:
        pedido.actualizar_dias_de_vencimiento()
    messages.success(request, 'Días de vencimiento actualizados para todos los pedidos.')
    return redirect('pedido_list_juan')


def actualizar_tasas_juan(request):
    pedidos = Pedido.objects.order_by('-id')[:50]
    for pedido in pedidos:
        pedido.actualizar_tasa_representativa()
    messages.success(request, 'Se Actualizo La TRM Con Banco De La Republica Correctamente')
    return redirect('pedido_list_juan')


# ----------------------------- Vista Autorización De Cancelaciones De Pedidos ---------------------------------------

# Solicitar Cancelacion Del pedido.
@login_required
def solicitar_cancelacion(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)

    if request.method == 'POST':
        observaciones = request.POST.get('observaciones', '')
        if pedido.estado_cancelacion == 'sin_solicitud' or pedido.estado_cancelacion == 'no_autorizado':
            autorizacion = AutorizacionCancelacion.objects.create(pedido=pedido, usuario_solicitante=request.user)
            pedido.estado_cancelacion = 'pendiente'
            pedido.observaciones = observaciones
            pedido.save()
        messages.warning(request, f'Se ha enviado la solicitud de cancelación para el pedido: {pedido.id}')
        return redirect('pedido_list_general')  # Redirigir a la lista de pedidos o a donde sea necesario

    return render(request, 'solicitar_cancelacion.html', {'pedido': pedido})


# Autorizar Cancelacion del pedido
@login_required
def autorizar_cancelacion(request, autorizacion_id):
    autorizacion = get_object_or_404(AutorizacionCancelacion, id=autorizacion_id)

    if request.method == 'POST':
        observaciones = request.POST.get('observaciones', '')
        accion = request.POST.get('accion', '')
        # Verificar si el usuario actual tiene permiso para autorizar la cancelación
        if request.user.groups.filter(name='Autorizadores').exists():
            pedido = autorizacion.pedido
            if accion == 'autorizar':
                autorizacion.autorizado = True
                autorizacion.fecha_autorizacion = timezone.now()
                autorizacion.usuario_autorizador = request.user
                autorizacion.save()
                pedido.estado_cancelacion = 'autorizado'
                pedido.estado_pedido = 'Cancelado'
                pedido.estado_factura = 'Cancelada'
                pedido.awb = '-'
                pedido.descuento = 0
                pedido.dias_de_vencimiento = 0
                pedido.diferencia_por_abono = 0
                pedido.documento_cobro_utilidad = 'Pedido Cancelado'
                pedido.estado_utilidad = 'Pedido Cancelado'
                pedido.numero_factura = 'Pedido Cancelado'
                pedido.tasa_representativa_usd_diaria = 0
                pedido.total_cajas_enviadas = 0
                pedido.total_peso_bruto = 0
                pedido.total_piezas_enviadas = 0
                pedido.total_piezas_solicitadas = 0
                pedido.trm_cotizacion = 0
                pedido.trm_monetizacion = 0
                pedido.utilidad_bancaria_usd = 0
                pedido.valor_total_factura_usd = 0
                pedido.valor_total_nota_credito_usd = 0
                pedido.valor_total_utilidad_usd = 0
                pedido.valor_utilidad_pesos = 0
            elif accion == 'no_autorizar':
                pedido.estado_cancelacion = 'no_autorizado'
                messages.info(request, f' Se ha anulado la cancelación para el pedido: {pedido.id}')
            pedido.observaciones = observaciones
            pedido.save()

            # Eliminar todos los detalles del pedido asociado si autorizado
            if accion == 'autorizar':
                DetallePedido.objects.filter(pedido=pedido).delete()
                messages.warning(request, f' Se ha cancelado correctamente el pedido: {pedido.id}')
        return redirect('pedido_list_general')
    return render(request, 'autorizar_cancelacion.html', {'autorizacion': autorizacion})


@login_required
def filtrar_presentaciones(request):
    fruta_id = request.GET.get('fruta_id')
    pedido_id = request.GET.get('pedido_id')

    if fruta_id and pedido_id:
        presentaciones = Presentacion.objects.filter(
            clientepresentacion__cliente__pedido__id=pedido_id,
            clientepresentacion__fruta_id=fruta_id
        ).values('id', 'nombre', 'kilos')
        return JsonResponse({'presentaciones': list(presentaciones)})
    return JsonResponse({'presentaciones': []})


@login_required
def load_referencias(request):
    presentacion_id = request.GET.get('presentacion_id')
    pedido_id = request.GET.get('pedido_id')
    tipo_caja_id = request.GET.get('tipo_caja_id')
    fruta_id = request.GET.get('fruta_id')

    if presentacion_id and tipo_caja_id and fruta_id and pedido_id:
        try:
            pedido = Pedido.objects.get(id=pedido_id)
            referencias = Referencias.objects.filter(
                presentacionreferencia__presentacion_id=presentacion_id,
                presentacionreferencia__tipo_caja_id=tipo_caja_id,
                presentacionreferencia__fruta_id=fruta_id,
                exportador=pedido.exportadora
            ).distinct()
        except Pedido.DoesNotExist:
            referencias = Referencias.objects.none()
    else:
        referencias = Referencias.objects.none()

    referencias_data = [{'id': ref.id, 'nombre': ref.nombre} for ref in referencias]
    return JsonResponse({'referencias': referencias_data})


# Exportacion de Resumen a PDF
@login_required(login_url=reverse_lazy('home'))
def export_pdf_resumen_semana(request):
    # Obtener los parámetros de filtro
    semana = request.GET.get('semana')
    exportador_id = request.GET.get('exportadora')

    # Verificar que los parámetros no estén vacíos
    if not semana:
        messages.error(request, 'Debe seleccionar por lo menos una semana.')
        return redirect('resumen_seguimiento_list_heavens')

    # Filtrar los datos basados en los parámetros
    pedidos = Pedido.objects.all()
    if semana:
        pedidos = pedidos.filter(semana=semana)
    if exportador_id:
        pedidos = pedidos.filter(exportadora_id=exportador_id)
        try:
            exportador_nombre = Exportador.objects.get(id=exportador_id).nombre
        except Exportador.DoesNotExist:
            exportador_nombre = None
    else:
        exportador_nombre = None

    # Generar la tabla con los datos filtrados
    table = SeguimienosResumenTable(pedidos)
    context = {
        'table': table,
        'semana': semana,
        'exportador': exportador_nombre
    }
    html_string = render_to_string('seguimiento_resumen_pdf.html', context, request=request)

    # Función para convertir HTML a PDF
    def convert_html_to_pdf(source_html, output_filename):
        result_file = open(output_filename, "w+b")
        pisa_status = pisa.CreatePDF(source_html, dest=result_file)
        result_file.close()
        return pisa_status.err

    # Generar el PDF
    result = convert_html_to_pdf(html_string, 'output.pdf')

    if result:
        return HttpResponse("Error al generar el PDF", status=500)

    with open('output.pdf', 'rb') as pdf:
        response = HttpResponse(pdf.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="seguimiento_resumen.pdf"'
        return response


@login_required(login_url=reverse_lazy('home'))
def exportar_pdf_resumen_pedido(request, pedido_id):
    start_time = time.time()

    # Obtener el pedido y los detalles del pedido
    pedido = get_object_or_404(Pedido, pk=pedido_id)
    if not request.user.groups.filter(name=pedido.exportadora.nombre).exists():
        return HttpResponseForbidden("No tienes permiso para ver estos detalles del pedido")
    step_time = time.time()
    print(f"Tiempo para verificar permisos y obtener el pedido: {step_time - start_time:.2f} segundos")

    # Filtrar los detalles del pedido
    detalles = DetallePedido.objects.filter(pedido=pedido).select_related('pedido')
    step_time_2 = time.time()
    print(f"Tiempo para filtrar los detalles del pedido: {step_time_2 - step_time:.2f} segundos")

    # Calcular totales
    total_cajas_solicitadas = detalles.aggregate(Sum('cajas_solicitadas'))['cajas_solicitadas__sum']
    total_peso_bruto = sum(detalle.calcular_peso_bruto() for detalle in detalles)
    total_piezas = math.ceil(sum(detalle.calcular_no_piezas() for detalle in detalles))
    step_time_3 = time.time()
    print(f"Tiempo para calcular los totales: {step_time_3 - step_time_2:.2f} segundos")

    # Generar la tabla con los datos filtrados
    table = ResumenPedidoTable(detalles)
    step_time_4 = time.time()
    print(f"Tiempo para generar la tabla: {step_time_4 - step_time_3:.2f} segundos")

    # Contexto para la plantilla
    context = {
        'pedido': pedido,
        'table': table,
        'total_cajas_solicitadas': total_cajas_solicitadas,
        'total_peso_bruto': total_peso_bruto,
        'total_piezas': total_piezas
    }

    # Renderizar la plantilla a HTML
    html_string = render_to_string('resumen_pedido_pdf.html', context, request=request)
    step_time_5 = time.time()
    print(f"Tiempo para renderizar la plantilla a HTML: {step_time_5 - step_time_4:.2f} segundos")

    # Función para convertir HTML a PDF
    def convert_html_to_pdf(source_html, output_filename):
        result_file = open(output_filename, "w+b")
        pisa_status = pisa.CreatePDF(source_html, dest=result_file)
        result_file.close()
        return pisa_status.err

    # Generar el PDF
    result = convert_html_to_pdf(html_string, 'output.pdf')
    step_time_6 = time.time()
    print(f"Tiempo para generar el PDF: {step_time_6 - step_time_5:.2f} segundos")

    if result:
        return HttpResponse("Error al generar el PDF", status=500)

    with open('output.pdf', 'rb') as pdf:
        response = HttpResponse(pdf.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="resumen_pedido_{pedido_id}.pdf"'
        end_time = time.time()
        print(f"Tiempo total: {end_time - start_time:.2f} segundos")
        return response


# -------------------- Vista para filtro de exportacion Seguimiento o Tracking -----------------------------------------


@login_required
@user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home'))
def exportar_excel_seguimiento_tracking(request):
    if request.method == 'POST':
        form = ExportSearchFormSeguimientos(request.POST)
        if form.is_valid():
            cliente = form.cleaned_data.get('cliente')
            intermediario = form.cleaned_data.get('intermediario')
            fecha_inicial = form.cleaned_data.get('fecha_inicial')
            fecha_final = form.cleaned_data.get('fecha_final')

            pedidos = Pedido.objects.all()

            if cliente:
                pedidos = pedidos.filter(cliente=cliente)
            if intermediario:
                pedidos = pedidos.filter(intermediario=intermediario)
            if fecha_inicial:
                pedidos = pedidos.filter(fecha_entrega__gte=fecha_inicial)
            if fecha_final:
                pedidos = pedidos.filter(fecha_entrega__lte=fecha_final)

            data = []
            for pedido in pedidos:
                data.append({
                    'Week': pedido.semana,
                    'Order No.': pedido.id,
                    'Request Date': pedido.fecha_solicitud.strftime('%d/%m/%Y') if pedido.fecha_solicitud else '',
                    'Exporter': pedido.exportadora.nombre if pedido.exportadora else '',
                    'Intermediary': pedido.intermediario.nombre if pedido.intermediario else '',
                    'Customer': pedido.cliente.nombre if pedido.cliente else '',
                    'Destination': pedido.destino.codigo if pedido.destino else '',
                    'Requested Boxes': pedido.total_cajas_solicitadas,
                    'Requested Pallets': pedido.total_piezas_solicitadas,
                    'Gross Weight': pedido.total_peso_bruto_solicitado,
                    'Delivery Date': pedido.fecha_entrega.strftime('%d/%m/%Y') if pedido.fecha_entrega else '',
                    'Awb': pedido.awb,
                    'Airline': pedido.aerolinea.nombre if pedido.aerolinea else '',
                    'Arrival Date': pedido.fecha_llegada.strftime('%d/%m/%Y') if pedido.fecha_llegada else '',
                    'Estimated Time of Arrival': pedido.eta.strftime('%d/%m/%Y') if pedido.eta else '',
                    'Estimated Time of Departure': pedido.etd.strftime('%d/%m/%Y') if pedido.etd else '',
                    'Cargo Agency': pedido.agencia_carga.nombre if pedido.agencia_carga else '',
                    'Products': pedido.variedades,
                    'Shipped Boxes': pedido.total_cajas_enviadas,
                    'Total Pallets Shipped': pedido.total_piezas_enviadas,
                    'Total Gross Weight Shipped	Weight': pedido.total_peso_bruto_enviado,
                    'Weight Awb': pedido.peso_awb,
                    'Kg Invoice / AWB': pedido.diferencia_peso_factura_awb,
                    'Real Final Eta': pedido.eta_real.strftime('%d/%m/%Y') if pedido.eta_real else '',
                    'Invoice': pedido.numero_factura,
                    'Termo': pedido.termo,
                    'Booking Responsible': pedido.responsable_reserva.nombre if pedido.responsable_reserva else '',
                    'Reserve Status': pedido.estatus_reserva,
                    'Document Status': pedido.estado_documentos,
                    'Order Status': pedido.estado_pedido,
                    'Comments': pedido.observaciones_tracking,
                })

            df = pd.DataFrame(data)

            wb = Workbook()
            ws = wb.active

            fill_red_soft = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        if row[df.columns.get_loc('Order Status')] == 'Cancelado':
                            cell.fill = fill_red_soft

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=seguimiento_pedidos.xlsx'
            wb.save(response)

            return response
    else:
        form = ExportSearchForm()

    return render(request, 'export_pedidos_tracking.html', {'form': form})


# -------------------- Exportacion Explicita de Excel Del resumen de exportaciones Por semana -------------------------

@login_required
@user_passes_test(es_miembro_del_grupo('Heavens'), login_url=reverse_lazy('home'))
def exportar_excel_seguimiento_resumen(request):
    semana = request.GET.get('semana')
    exportador_id = request.GET.get('exportadora')

    # Verificar que los parámetros no estén vacíos
    if not semana:
        messages.error(request, 'Debe seleccionar por lo menos una semana.')
        return redirect('resumen_seguimiento_list_heavens')

    # Filtrar los datos basados en los parámetros
    pedidos = Pedido.objects.all()
    if semana:
        pedidos = pedidos.filter(semana=semana)
    if exportador_id:
        pedidos = pedidos.filter(exportadora_id=exportador_id)
        try:
            exportador_nombre = Exportador.objects.get(id=exportador_id).nombre
        except Exportador.DoesNotExist:
            exportador_nombre = "Unknown"
    else:
        exportador_nombre = "Unknown"

    # Preparar los datos para el DataFrame
    data = []
    for pedido in pedidos:
        data.append({
            'Week': pedido.semana,
            'Order No.': pedido.id,
            'Exporter': pedido.exportadora.nombre if pedido.exportadora else '',
            'Customer': pedido.cliente.nombre if pedido.cliente else '',
            'Destination': pedido.destino.codigo if pedido.destino else '',
            'Products': pedido.variedades,
            'Requested Boxes': pedido.total_cajas_solicitadas,
            'Shipped Boxes': pedido.total_cajas_enviadas,
            'Requested Pallets': pedido.total_piezas_solicitadas,
            'Delivery Date': pedido.fecha_entrega.strftime('%d/%m/%Y') if pedido.fecha_entrega else '',
            'Booking Responsible': pedido.responsable_reserva.nombre if pedido.responsable_reserva else '',
            'Reserve Status': pedido.estatus_reserva,
            'Airline': pedido.aerolinea.nombre if pedido.aerolinea else '',
            'Cargo Agency': pedido.agencia_carga.nombre if pedido.agencia_carga else '',
            'Order Status': pedido.estado_pedido,
            'Document Status': pedido.estado_documentos,
            'Tracking Comments': pedido.observaciones_tracking,
        })

    df = pd.DataFrame(data)

    # Crear el archivo Excel con estilo
    wb = Workbook()
    ws = wb.active

    # Definir el color de relleno rojo suave
    fill_red_soft = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Agregar el DataFrame al archivo Excel
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Si es el encabezado
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                # Pintar la fila si el estado del pedido es 'cancelado'
                if row[df.columns.get_loc('Order Status')] == 'Cancelado':
                    cell.fill = fill_red_soft

    # Ajustar el ancho de las columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=summary_week_{semana}_exporter_{exportador_nombre}.xlsx'
    wb.save(response)

    return response
